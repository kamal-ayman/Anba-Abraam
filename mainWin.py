from openpyxl import load_workbook

from time import sleep
import os
import MySQLdb
from PyQt5.QtCore import *
from PyQt5.QtWidgets import *
from PyQt5.QtWidgets import QMainWindow, QApplication

import pics_rc
from index import Ui_MainWindow as FORM


class Main(QMainWindow, FORM):
    def __init__(self, parent=None):
        super(Main, self).__init__(parent)
        QMainWindow.__init__(self)
        self.setupUi(self)
        self.TryPassword = 0
        self.InItUI()
        self.DB_Connect()
        self.UpdateNumberUsers()
        self.SignUpTab()
        self.Buttons()
        self.ShowComboBox()

    def LogoText(self, text):
        for tub in text:
            self.label_26.setText(tub)

    def StatusBar(self, text):
        self.statusbar.showMessage(text)

    def ShowComboBox(self):
        self.ShowFatherOfConfession()
        self.ShowHomeContentComboBox()
        self.ShowGovernorate()
        self.ShowDistrict()
        self.ShowAddress()
        self.ShowSource()

    def InItUI(self):
        self.tabWidget.tabBar().setVisible(False)
        self.tabWidget_2.tabBar().setVisible(False)
        self.tabWidget_3.tabBar().setVisible(False)
        self.tabWidget_4.tabBar().setVisible(False)
        self.tabWidget_5.tabBar().setVisible(False)
        self.tabWidget_6.tabBar().setVisible(False)
        self.tabWidget_7.tabBar().setVisible(False)
        self.progressBar.setVisible(False)

    def DB_Connect(self):
        self.db = MySQLdb.connect(host='127.0.0.1', user='root', password='toor', port=3306, db='anbaabram',
                                  charset='utf8', init_command='SET NAMES UTF8')
        self.cur = self.db.cursor()

    def HighMainButton(self):
        self.pushButton_3.setHidden(True)
        self.pushButton.setHidden(True)
        self.pushButton_2.setHidden(True)
        # -- Reset Password -- #
        self.pushButton_43.setHidden(True)

    def ShowMainButton(self):
        self.pushButton_3.setHidden(False)
        self.pushButton.setHidden(False)
        self.pushButton_2.setHidden(False)

    def Buttons(self):
        # -- Open Main Tabs -- #
        self.pushButton_3.clicked.connect(self.OpenShowTab)
        self.pushButton.clicked.connect(self.OpenAdds)
        self.pushButton_2.clicked.connect(self.OpenEditOrDeleteTab)
        # -- Sign Up Start Program -- #
        self.pushButton_83.clicked.connect(self.SignUpTab)
        # -- Sign Up -- #
        self.pushButton_44.clicked.connect(self.SignUpAddAccount)
        # -- Login -- #
        self.pushButton_42.clicked.connect(self.Login)
        # ----------------------------------------------------------------- #
        #                     -- Click Add Next --                          #
        # ----------------------------------------------------------------- #
        self.pushButton_5.clicked.connect(self.AddNext0)
        self.pushButton_7.clicked.connect(self.AddNext1)
        self.pushButton_9.clicked.connect(self.AddNext2_1)
        self.pushButton_13.clicked.connect(self.AddNext2_2)
        self.pushButton_15.clicked.connect(self.AddNext3_1)
        self.pushButton_17.clicked.connect(self.AddNext3_2)
        self.pushButton_25.clicked.connect(self.AddNext4)
        # ----------------------------------------------------------------- #
        #                     -- Click Add Back --                          #
        # ----------------------------------------------------------------- #
        self.pushButton_6.clicked.connect(self.AddBack1)
        self.pushButton_8.clicked.connect(self.AddBack2_1)
        self.pushButton_12.clicked.connect(self.AddBack2_2)
        self.pushButton_14.clicked.connect(self.AddBack3_1)
        self.pushButton_16.clicked.connect(self.AddBack3_2)
        self.pushButton_18.clicked.connect(self.AddBack4)
        self.pushButton_54.clicked.connect(self.AddBack5)
        # ----------------------------------------------------------------- #
        #                     -- Click Edit Next --                         #
        # ----------------------------------------------------------------- #
        self.pushButton_27.clicked.connect(self.EditNext0)
        self.pushButton_29.clicked.connect(self.EditNext1)
        self.pushButton_36.clicked.connect(self.EditNext2_1)
        self.pushButton_38.clicked.connect(self.EditNext2_2)
        self.pushButton_41.clicked.connect(self.EditNext3_1)
        self.pushButton_48.clicked.connect(self.EditNext3_2)
        self.pushButton_53.clicked.connect(self.EditNext4)
        # ----------------------------------------------------------------- #
        #                     -- Click Edit Back --                         #
        # ----------------------------------------------------------------- #
        self.pushButton_28.clicked.connect(self.EditBack1)
        self.pushButton_35.clicked.connect(self.EditBack2_1)
        self.pushButton_39.clicked.connect(self.EditBack2_2)
        self.pushButton_40.clicked.connect(self.EditBack3_1)
        self.pushButton_49.clicked.connect(self.EditBack3_2)
        self.pushButton_50.clicked.connect(self.EditBack4)
        self.pushButton_51.clicked.connect(self.EditBack5)
        # ----------------------------------------------------------------- #
        #                     -- FatherOfConfession Add --                  #
        # ----------------------------------------------------------------- #
        self.pushButton_66.clicked.connect(self.AddFatherOfConfession)
        self.pushButton_65.clicked.connect(self.DeleteFatherOfConfession)
        # ----------------------------------------------------------------- #
        self.pushButton_19.clicked.connect(self.AddFatherOfConfession)
        self.pushButton_60.clicked.connect(self.DeleteFatherOfConfession)
        # ----------------------------------------------------------------- #
        self.pushButton_68.clicked.connect(self.AddFatherOfConfession)
        self.pushButton_67.clicked.connect(self.DeleteFatherOfConfession)
        # ----------------------------------------------------------------- #
        #                    -- FatherOfConfession Edit --                  #
        # ----------------------------------------------------------------- #
        self.pushButton_63.clicked.connect(self.AddFatherOfConfession)
        self.pushButton_64.clicked.connect(self.DeleteFatherOfConfession)
        # ----------------------------------------------------------------- #
        self.pushButton_62.clicked.connect(self.AddFatherOfConfession)
        self.pushButton_61.clicked.connect(self.DeleteFatherOfConfession)
        # ----------------------------------------------------------------- #
        self.pushButton_73.clicked.connect(self.AddFatherOfConfession)
        self.pushButton_74.clicked.connect(self.DeleteFatherOfConfession)
        # ----------------------------------------------------------------- #
        #                        -- Governorate Add --                      #
        # ----------------------------------------------------------------- #
        self.pushButton_20.clicked.connect(self.AddGovernorate)
        self.pushButton_22.clicked.connect(self.DeleteGovernorate)
        # ----------------------------------------------------------------- #
        #                        -- Governorate Edit --                     #
        # ----------------------------------------------------------------- #
        self.pushButton_113.clicked.connect(self.AddGovernorate)
        self.pushButton_116.clicked.connect(self.DeleteGovernorate)
        # ----------------------------------------------------------------- #
        #                        -- District Add --                         #
        # ----------------------------------------------------------------- #
        self.pushButton_21.clicked.connect(self.AddDistrict)
        self.pushButton_23.clicked.connect(self.DeleteDistrict)
        # ----------------------------------------------------------------- #
        #                        -- District Edit --                        #
        # ----------------------------------------------------------------- #
        self.pushButton_115.clicked.connect(self.AddDistrict)
        self.pushButton_33.clicked.connect(self.DeleteDistrict)
        # ----------------------------------------------------------------- #
        #                        -- Address Add --                          #
        # ----------------------------------------------------------------- #
        self.pushButton_76.clicked.connect(self.AddAddress)
        self.pushButton_75.clicked.connect(self.DeleteAddress)
        # ----------------------------------------------------------------- #
        self.pushButton_114.clicked.connect(self.AddAddress)
        self.pushButton_117.clicked.connect(self.DeleteAddress)
        # ----------------------------------------------------------------- #
        #                        -- Source Add --                           #
        # ----------------------------------------------------------------- #
        self.pushButton_70.clicked.connect(self.AddSource)
        self.pushButton_69.clicked.connect(self.DeleteSource)
        # ----------------------------------------------------------------- #
        #                        -- Source Edit --                          #
        # ----------------------------------------------------------------- #
        self.pushButton_71.clicked.connect(self.AddSource)
        self.pushButton_72.clicked.connect(self.DeleteSource)
        # ----------------------------------------------------------------- #
        #                           -- Home Contents --                     #
        # ----------------------------------------------------------------- #
        self.pushButton_80.clicked.connect(self.UpdateHomeContentComboBox)
        self.pushButton_125.clicked.connect(self.DeleteHomeContent)
        # ----------------------------------------------------------------- #
        self.pushButton_82.clicked.connect(self.UpdateHomeContentComboBox)
        self.pushButton_126.clicked.connect(self.DeleteHomeContent)
        # ----------------------------------------------------------------- #
        self.pushButton_79.clicked.connect(self.AddAddHomeContents)
        self.pushButton_85.clicked.connect(self.AddRemoveHomeContents)
        # ----------------------------------------------------------------- #
        self.pushButton_81.clicked.connect(self.EditAddHomeContents)
        self.pushButton_86.clicked.connect(self.EditRemoveHomeContents)
        # ----------------------------------------------------------------- #
        # ----------------------------------------------------------------- #
        self.pushButton_24.clicked.connect(self.OpenFile)
        self.pushButton_26.clicked.connect(self.PrintFile)
        # ----------------------------------------------------------------- #
        # ----------------------------------------------------------------- #
        #                                                                   #
        #                                                                   #
        #                            -- Add Tab --                          #
        #                                                                   #
        #                                                                   #
        # ----------------------------------------------------------------- #
        # ----------------------------------------------------------------- #
        #                             -- Add Data --                        #

        self.pushButton_77.clicked.connect(self.AddData)
        # ----------------------------------------------------------------- #
        #                         --  Add Children --                       #

        self.pushButton_10.clicked.connect(self.AddChildren)
        # ----------------------------------------------------------------- #
        #                           -- Add Human --                         #

        self.pushButton_11.clicked.connect(self.AddHumanData)
        # ----------------------------------------------------------------- #
        # ----------------------------------------------------------------- #
        #                                                                   #
        #                                                                   #
        #                            -- Edit Tab --                         #
        #                                                                   #
        #                                                                   #
        # ----------------------------------------------------------------- #
        # ----------------------------------------------------------------- #
        #                        -- Edit Search Data --                     #
        # ----------------------------------------------------------------- #
        self.pushButton_52.clicked.connect(self.SearchEdit)
        # ----------------------------------------------------------------- #
        #                   -- Update Husband And Wife Data --              #
        # ----------------------------------------------------------------- #
        self.pushButton_84.clicked.connect(self.UpdateHusbandAndWifeData)
        # ----------------------------------------------------------------- #
        #                        -- Update Info Data --                     #
        # ----------------------------------------------------------------- #
        self.pushButton_118.clicked.connect(self.SaveInfo)
        # ----------------------------------------------------------------- #
        #                        -- Update Info Data --                     #
        # ----------------------------------------------------------------- #
        self.pushButton_59.clicked.connect(self.UpdateOtherInfoData)
        # ----------------------------------------------------------------- #
        #                        --  Save Last Page --                      #
        # ----------------------------------------------------------------- #
        self.pushButton_131.clicked.connect(self.SaveLastPage)
        # ----------------------------------------------------------------- #
        #                     --  Save Last Page Part2 --                   #
        # ----------------------------------------------------------------- #
        self.pushButton_132.clicked.connect(self.SaveLastPagePart2)
        # ----------------------------------------------------------------- #
        #                        -- Delete All Data --                      #
        # ----------------------------------------------------------------- #
        self.pushButton_130.clicked.connect(self.DeleteAllData)
        # ----------------------------------------------------------------- #
        #                        --  Edit Children --                       #
        # ----------------------------------------------------------------- #
        self.pushButton_56.clicked.connect(self.SearchChildren)
        self.pushButton_34.clicked.connect(self.UpdateChildren)
        self.pushButton_57.clicked.connect(self.DeleteChildrenData)
        # ----------------------------------------------------------------- #
        #                          -- Edit Human --                         #
        # ----------------------------------------------------------------- #
        self.pushButton_119.clicked.connect(self.SearchHuman)
        self.pushButton_37.clicked.connect(self.UpdateHuman)
        self.pushButton_58.clicked.connect(self.DeleteHumanData)

    # ----------------------------------------------------------------- #
    #                          -- Start Prog  --                        #
    # ----------------------------------------------------------------- #

    def SignUpTab(self):
        self.HighMainButton()
        self.cur.execute("""SELECT * FROM anbaabram.users""")
        Data = self.cur.fetchall()
        if Data == ((1, ' ', ' ', None, 0),):
            self.BG()
            self.tabWidget.setCurrentIndex(0)
        else:
            self.worker = Process()
            self.worker.start()
            self.worker.T.connect(self.LogoText)
            self.BG1()
            self.tabWidget.setCurrentIndex(1)

    def UpdateNumberUsers(self):
        self.cur.execute(f"""SELECT NumberOfUsers FROM anbaabram.users """)
        NumOfUsers = self.cur.fetchall()
        n = []
        for num in NumOfUsers:
            n.append(num[0])
        self.n = int(n[-1]) + 1
        self.lcdNumber.setProperty("intValue", int(n[-1]))

    def UpdateNumberChildren(self):
        try:
            global NumChildren
            NumFamily = self.lineEdit_16.text()
            self.cur.execute(f"""SELECT ChildrenNumber FROM anbaabram.children WHERE NumberFamily = "{NumFamily}" """)
            Data = self.cur.fetchall()
            for data in Data:
                NumChildren = int(data[0])
            self.lcdNumber_2.setProperty("intValue", int(NumChildren))
        except:
            pass

    def UpdateNumberOtherHuman(self):
        try:
            global HumanNum
            NumFamily = self.lineEdit_16.text()
            self.cur.execute(f"""SELECT HumanNum FROM anbaabram.anotherhuman WHERE NumberFamily = "{NumFamily}" """)
            Data = self.cur.fetchall()
            for data in Data:
                HumanNum = int(data[0])
            self.lcdNumber_3.setProperty("intValue", int(HumanNum))
        except:
            pass

    def SignUpAddAccount(self):
        UserName = self.lineEdit_40.text()
        Pass1 = self.lineEdit_39.text()
        Pass2 = self.lineEdit_112.text()
        Email = self.lineEdit_41.text()
        if UserName != "" and Pass1 != "" and Pass1 == Pass2:
            try:
                self.cur.execute(
                    f"""INSERT INTO anbaabram.users (UserName, Password, Email, NumberOfUsers) VALUES ("{UserName}", "{Pass1}", "{Email}", "{self.n}") """)
                self.db.commit()
                self.UpdateNumberUsers()
                self.lineEdit_40.setText("")
                self.lineEdit_39.setText("")
                self.lineEdit_112.setText("")
                self.lineEdit_41.setText("")
                self.statusbar.showMessage("كمستخدم جديد -- {} -- تم اضافة".format(UserName))
                self.worker = ClearStatusBar()
                self.worker.start()
                self.worker.ClearStatusBarVariable.connect(self.StatusBar)

            except:
                QMessageBox.about(self, "Error", "اسم المستخدم موجود بالفعل !!")
        else:
            QMessageBox.warning(self, "Error", "الرجاء ادخال جميع البيانات للتسجيل")

    def Login(self):
        UserName = self.lineEdit.text()
        Password = self.lineEdit_1.text()
        if UserName == "" and Password == "":
            QMessageBox.warning(self, "Error", "الرجاء ادخال اسم المستخدم وكلمة السر")
        elif UserName == "":
            QMessageBox.warning(self, "Error", "الرجاء ادخال اسم المستخدم ...")
        elif Password == "":
            QMessageBox.warning(self, "Error", "الرجاء ادخال كلمة السر ...")
        else:
            self.cur.execute(
                f"""SELECT * FROM anbaabram.users WHERE UserName = "{UserName}" AND Password = "{Password}" """)
            Data = self.cur.fetchall()
            if Data != ():
                self.tabWidget.setCurrentIndex(3)
                self.MoveLogo()
                self.ShowMainButton()
            else:
                if self.TryPassword < 3:
                    if self.TryPassword == 2:
                        self.pushButton_43.setHidden(False)
                    self.TryPassword += 1
                    QMessageBox.warning(self, "Error", "خطأ ف اسم المستخدم او كلمة المرور !")
                else:
                    QMessageBox.warning(self, "Error", "خطأ ف اسم المستخدم او كلمة المرور !")

    # -------------------------------------------------------------- #
    #                   -- Show, Open, PrintFile --                  #
    # -------------------------------------------------------------- #

    def ReturnValues(self):
        NumFamily = ""
        Status = True
        HusbandStatus, WifeStatus = True, True
        LineText = self.lineEdit_38.text()
        combo = self.comboBox_21.currentIndex()

        if combo == 0:
            if LineText != "":
                NumFamily = self.lineEdit_38.text()
            else:
                Status = False
                QMessageBox.warning(self, "Error", "برجاء ادخال رقم الاسرة اولا ...")

        if combo == 1:
            if LineText != "":
                try:
                    self.cur.execute(
                        f"""SELECT NumberFamily FROM anbaabram.wifedata WHERE WifeNationalId = "{LineText}" """)
                    NumFamily = self.cur.fetchone()[0]
                except:
                    WifeStatus = False

                try:
                    self.cur.execute(
                        f"""SELECT NumberFamily FROM anbaabram.husbanddata WHERE HusbandNationalId = "{LineText}" """)
                    NumFamily = self.cur.fetchone()[0]
                except:
                    HusbandStatus = False
                if WifeStatus or HusbandStatus:
                    Status = True
                else:
                    Status = False
                    QMessageBox.warning(self, "Error", "الرقم القومي غير صحيح !")
            else:
                Status = False
                QMessageBox.warning(self, "Error", "برجاء ادخال الرقم القومي اولا ...")

        if Status:
            HusbandStatus, WifeStatus = True, True
            filename = "file.xlsx"
            # Make file
            import MK
            MK.MakeExFile()
            wb = load_workbook(filename)
            ws = wb.worksheets[0]
            self.cur.execute(
                f"""SELECT HusbandName, HusbandNikeName, HusbandNationalId, HusbandWork, HusbandMoney, HusbandPhoneNumber, HusbandFatherOfConfession_id FROM anbaabram.husbanddata WHERE NumberFamily = "{NumFamily}" """)
            HusbandData = self.cur.fetchone()
            try:
                if HusbandData is not None:
                    num = 16
                    for n, item in enumerate(HusbandData):
                        print(HusbandData[n])
                        ws[f'B{num + n}'] = HusbandData[n]
                else:
                    HusbandStatus = False
            except:
                pass

            self.cur.execute(
                f"""SELECT WifeName, WifeNikeName, WifeNationalId, WifeWork, WifeMoney, WifePhoneNumber, WifeFatherOfConfession_id FROM anbaabram.wifedata WHERE NumberFamily = "{NumFamily}" """)
            WifeData = self.cur.fetchone()
            try:
                if WifeData is not None:
                    num = 16
                    for n, item in enumerate(WifeData):
                        ws[f'E{num + n}'] = WifeData[n]
                else:
                    WifeStatus = False
            except:
                pass

            if HusbandStatus or WifeStatus:
                ws['E10'] = NumFamily
                self.cur.execute(
                    f"""SELECT DetailedAddress, Governorate, neighborhood, Region, SpecialMarque, Living, ChurchName, Address, Phone1, Phone2 FROM anbaabram.information WHERE NumFamily = "{NumFamily}" """)
                InfoData = self.cur.fetchone()
                print(InfoData)
                num = 24
                for n, item in enumerate(InfoData):
                    ws[f'B{num + n}'] = InfoData[n]

                abc = ['A', 'B', 'C', 'D', 'E', 'F']

                try:
                    self.cur.execute(
                        f"""SELECT ChildrenNumber FROM anbaabram.children WHERE NumberFamily = "{NumFamily}" """)
                    NumChildren = self.cur.fetchall()[-1][0]
                    print(NumChildren)
                    ws['B38'] = NumChildren
                    self.cur.execute(
                        f"""SELECT ChildrenName, ChildrenNationalId, ChildrenSocialStatus, ChildrenSchoolWork, ChildrenMonthlyIncome, ChildrenFatherOfConfession FROM anbaabram.children WHERE NumberFamily = "{NumFamily}" """)
                    ChildrenData = self.cur.fetchall()
                    print(ChildrenData)
                    for N, data in enumerate(ChildrenData):
                        for n, item in enumerate(data):
                            ws[f'{abc[n]}{N + 41}'] = item
                except:
                    pass
                try:
                    self.cur.execute(
                        f"""SELECT HumanNum FROM anbaabram.anotherhuman WHERE NumberFamily = "{NumFamily}" """)
                    NumOtherHuman = self.cur.fetchall()[-1][0]
                    ws['D47'] = NumOtherHuman
                    self.cur.execute(
                        f"""SELECT HumanName, HumanNationalId, HumanRelativeRelation, HumanSocialStatus, HumanFatherOfConfession, HumanMonthlyIncome FROM anbaabram.anotherhuman WHERE NumberFamily = "{NumFamily}" """)
                    OtherHuman = self.cur.fetchall()
                    print(OtherHuman)
                    for N, data in enumerate(OtherHuman):
                        for n, item in enumerate(data):
                            ws[f'{abc[n]}{N + 50}'] = item
                except:
                    pass
                try:
                    self.cur.execute(
                        f"""SELECT HomeContentsName, HomeContentsNumber FROM anbaabram.hcontents WHERE FamilyNum = "{NumFamily}" """)
                    HomeContentsData = self.cur.fetchall()
                    for N, data in enumerate(HomeContentsData):
                        for n, item in enumerate(data):
                            ws[f'{abc[n + 1]}{N + 56}'] = item
                except:
                    pass

                try:
                    self.cur.execute(
                        f"""SELECT Diseases, Obstruction, OtherCircumstances FROM anbaabram.anotherinfofamily WHERE NumFamily = "{NumFamily}" """)
                    OtherInfoData = self.cur.fetchone()
                    # for n, item in enumerate(OtherInfoData):
                    print(OtherInfoData)
                    num = 69
                    for n, item in enumerate(OtherInfoData):
                        ws[f'B{num}'] = item
                        num = num + 3
                except:
                    pass

                try:
                    self.cur.execute(
                        f"""SELECT SourceOfIncome, MonthlyIncome FROM anbaabram.primarysourcesofincome WHERE NumFamily = "{NumFamily}" """)
                    IncomeData = self.cur.fetchone()
                    ws['A81'] = IncomeData[0]
                    ws['B81'] = IncomeData[1]
                except:
                    pass

                try:
                    self.cur.execute(
                        f"""SELECT MonthlyMaterialAssistance, CurativeHelp, MonthlyGood, OtherAid, Others FROM anbaabram.churchesaidandgoods WHERE NumFamily = "{NumFamily}" """)
                    ChurchEsaidAndGoods = self.cur.fetchone()
                    ChurchEsaidAndGoods = list(ChurchEsaidAndGoods)
                    for n in range(4):
                        if ChurchEsaidAndGoods[n] == '1':
                            ChurchEsaidAndGoods[n] = 'نعم'
                        else:
                            ChurchEsaidAndGoods[n] = 'لا'
                        print(ChurchEsaidAndGoods)
                    ws['B86'] = ChurchEsaidAndGoods[0]
                    ws['E86'] = ChurchEsaidAndGoods[1]
                    ws['B88'] = ChurchEsaidAndGoods[2]
                    ws['E88'] = ChurchEsaidAndGoods[3]
                    ws['B90'] = ChurchEsaidAndGoods[4]
                except:
                    pass
                try:
                    self.cur.execute(
                        f"""SELECT * FROM anbaabram.churchesaidandgoodspart2 WHERE NumFamily = "{FamilyNum}" """)
                    ChurchEsaidAndGoodspart2 = self.cur.fetchone()
                    ChurchEsaidAndGoodspart2 = list(ChurchEsaidAndGoodspart2)
                    print(ChurchEsaidAndGoodspart2)
                    for n in range(6):
                        ws[f'B{96 + n}'] = ChurchEsaidAndGoodspart2[1 + n]
                    for n in range(6):
                        ws[f'E{96 + n}'] = ChurchEsaidAndGoodspart2[7 + n]
                except:
                    pass
                wb.save(filename)
                return True
            else:
                QMessageBox.warning(self, "Error", "رقم اسرة غير صحيح !")
                return False

    def OpenFile(self):
        try:
            if self.ReturnValues():
                os.startfile("file.xlsx", "open")
        except Exception as n:
            print(n)
            question = QMessageBox.question(self, "Error",
                                            " لا يمكنك البحث عندما يعمل ملف البحث لا يمكن تحديث البيانات \n هل تريد غلق ملف البحث لأستكمال البحث الجديد؟",
                                            QMessageBox.No | QMessageBox.Close)
            if question == QMessageBox.Close:
                os.system("TASKKILL /F /IM EXCEL.EXE")

    def PrintFile(self):
        try:
            if self.ReturnValues():
                os.startfile("file.xlsx", "print")
        except Exception as n:
            print(n)
            question = QMessageBox.question(self, "Error",
                                            " لا يمكنك البحث عندما يعمل ملف البحث لا يمكن تحديث البيانات \n هل تريد غلق ملف البحث لأستكمال البحث الجديد؟",
                                            QMessageBox.No | QMessageBox.Close)
            if question == QMessageBox.Close:
                os.system("TASKKILL /F /IM EXCEL.EXE")

    # ----------------------------------------------------------------- #
    # ----------------------------------------------------------------- #
    #                                                                   #
    #                          -- Add Tab --                            #
    #                                                                   #
    # ----------------------------------------------------------------- #
    # ----------------------------------------------------------------- #

    # ------------------------------------------------------------- #
    #                     -- Add All New Data --                    #
    #                                                               #
    def AddData(self):
        # -- number family -- # ----------------------------------------------
        global ChurchName, Living
        NumFamily = self.lineEdit_16.text()
        # -- Husband Data -- # ----------------------------------------------
        HusbandName = self.lineEdit_2.text()
        HusbandNikName = self.lineEdit_3.text()
        HusbandId = self.lineEdit_4.text()
        HusbandWork = self.lineEdit_5.text()
        HusbandMoney = self.lineEdit_6.text()
        HusbandPhone = self.lineEdit_7.text()
        HusbandFatherOfConfession = self.comboBox_20.currentText()
        HusbandStatus = True
        # -- Wife Data -- # ---------------------------------------------------
        WifeName = self.lineEdit_12.text()
        WifeNikeName = self.lineEdit_13.text()
        WifeId = self.lineEdit_9.text()
        WifeWork = self.lineEdit_14.text()
        WifeMoney = self.lineEdit_15.text()
        WifePhone = self.lineEdit_11.text()
        WifeFatherOfConfession = self.comboBox_19.currentText()
        WifeStatus = True
        # -- Info -- # ----------------------------------------------------------
        DetailedAddress = self.textEdit.toPlainText()
        Governorate = self.comboBox.currentText()
        neighborhood = self.comboBox_2.currentText()
        Region = self.lineEdit_19.text()
        SpecialMarque = self.lineEdit_20.text()
        if self.radioButton.isChecked():
            Living = self.radioButton.text()
        elif self.radioButton_2.isChecked():
            Living = self.radioButton_2.text()
        elif self.radioButton_4.isChecked():
            Living = self.radioButton_4.text()
        elif self.radioButton_3.isChecked():
            Living = self.radioButton_3.text()
        else:
            QMessageBox.warning(self, "Error", "تحقق من نظام السكن ...")
        if self.radioButton_5.isChecked() is True:
            ChurchName = self.radioButton_5.text()
        elif self.radioButton_6.isChecked() is True:
            ChurchName = self.lineEdit_21.text()
        else:
            QMessageBox.warning(self, "Error", "تحقق من اسم الكنيسة ...")
        Address = self.comboBox_3.currentText()
        Phone1 = self.lineEdit_24.text()
        Phone2 = self.lineEdit_23.text()
        # ------------------------------------------------------
        Diseases = self.textEdit_2.toPlainText()
        Obstruction = self.textEdit_3.toPlainText()
        OtherCircumstances = self.textEdit_4.toPlainText()
        # -----------------------------------------------------
        Source = self.comboBox_6.currentText()
        MonthlyIncome = self.lineEdit_37.text()
        # -----------------------------------------------------
        ChurchesAidAndGoods = self.lineEdit_58.text()
        TherapeuticAid = self.lineEdit_59.text()
        AidDuringTheStudy = self.lineEdit_60.text()
        BasicSalary = self.lineEdit_61.text()
        AdditionalResource_project = self.lineEdit_83.text()

        ElectricityWaterAndGas = self.lineEdit_62.text()
        Phone = self.lineEdit_63.text()
        Rent = self.lineEdit_64.text()
        Therapy = self.lineEdit_65.text()
        AStudy = self.lineEdit_84.text()

        # --------------------------------------------
        MonthlyMaterialAssistance = self.checkBox_3.checkState().__str__()
        CurativeHelp = self.checkBox_5.checkState().__str__()
        MonthlyGood = self.checkBox_4.checkState().__str__()
        OtherAid = self.checkBox_6.checkState().__str__()
        S_H1 = "0"
        S_H3 = "0"
        S_H2 = "0"
        S_H4 = "0"
        if MonthlyMaterialAssistance == "2":
            S_H1 = "1"
        if CurativeHelp == "2":
            S_H2 = "1"
        if MonthlyGood == "2":
            S_H3 = "1"
        if OtherAid == "2":
            S_H4 = "1"

        Others = self.textEdit_5.toPlainText()

        try:
            if HusbandName != "":
                try:
                    self.cur.execute(
                        f"""INSERT INTO anbaabram.husbanddata (HusbandName, HusbandNikeName, HusbandNationalId, HusbandWork, HusbandMoney, HusbandPhoneNumber, HusbandFatherOfConfession_id, NumberFamily) VALUES ("{HusbandName}", "{HusbandNikName}", "{HusbandId}", "{HusbandWork}", "{HusbandMoney}", "{HusbandPhone}", "{HusbandFatherOfConfession}", "{NumFamily}") """)
                    self.db.commit()
                except:
                    HusbandStatus = False
                    QMessageBox.warning(self, "Error", "رقم الاسرة او الرقم القومي للزوج موجود بالفعل...!")
            if WifeName != "":
                try:
                    self.cur.execute(
                        f"""INSERT INTO anbaabram.wifedata (WifeName, WifeNikeName, WifeNationalId, WifeWork, WifeMoney, WifePhoneNumber, WifeFatherOfConfession_id, NumberFamily) VALUES ("{WifeName}", "{WifeNikeName}", "{WifeId}", "{WifeWork}", "{WifeMoney}", "{WifePhone}", "{WifeFatherOfConfession}", "{NumFamily}") """)
                    self.db.commit()
                except:
                    WifeStatus = False
                    QMessageBox.warning(self, "Error", "رقم الاسرة او الرقم القومي للزوجة موجود بالفعل...!")

            if HusbandStatus == True and WifeStatus == True:
                try:
                    Total1 = float(ChurchesAidAndGoods) + float(TherapeuticAid) + float(AidDuringTheStudy) + float(
                        BasicSalary) + float(AdditionalResource_project)
                    Total2 = float(ElectricityWaterAndGas) + float(Phone) + float(Rent) + float(Therapy) + float(AStudy)

                    self.cur.execute(f"""INSERT INTO anbaabram.churchesaidandgoodspart2 
                    (ChurchesAidAndGoods, TherapeuticAid, AidDuringTheStudy, BasicSalary, AdditionalResource_project, Total1, 
                    ElectricityWaterAndGas, Phone, Rent, Therapy, AStudy, Total2, NumFamily) 
                    VALUES ("{ChurchesAidAndGoods}", "{TherapeuticAid}", "{AidDuringTheStudy}", "{BasicSalary}", "{AdditionalResource_project}", "{Total1}", 
                    "{ElectricityWaterAndGas}", "{Phone}", "{Rent}", "{Therapy}", "{AStudy}", "{Total2}", "{NumFamily}") """)
                    self.db.commit()

                    try:
                        self.cur.execute(
                            f"""INSERT INTO anbaabram.information (DetailedAddress, Governorate, neighborhood, Region, SpecialMarque, Living, ChurchName, Address, Phone1, Phone2, NumFamily) VALUES ("{DetailedAddress}", "{Governorate}", "{neighborhood}", "{Region}", "{SpecialMarque}", "{Living}", "{ChurchName}", "{Address}", "{Phone1}", "{Phone2}", "{NumFamily}") """)
                        self.db.commit()
                        try:
                            self.cur.execute(
                                f"""INSERT INTO anbaabram.anotherinfofamily (Diseases, Obstruction, OtherCircumstances, NumFamily) VALUES ("{Diseases}", "{Obstruction}", "{OtherCircumstances}", "{NumFamily}") """)
                            self.db.commit()
                            try:
                                self.cur.execute(
                                    f"""INSERT INTO anbaabram.churchesaidandgoods (MonthlyMaterialAssistance, CurativeHelp, MonthlyGood, OtherAid, Others, NumFamily) VALUES ("{S_H1}", "{S_H2}", "{S_H3}", "{S_H4}", "{Others}", "{NumFamily}") """)
                                self.db.commit()
                                try:
                                    self.cur.execute(
                                        f"""INSERT INTO anbaabram.primarysourcesofincome (SourceOfIncome, MonthlyIncome, NumFamily) VALUES ("{Source}", "{MonthlyIncome}", "{NumFamily}") """)
                                    self.db.commit()
                                    self.statusbar.showMessage("تم حفظ البايانات بنجاح")
                                    self.tabWidget_2.setCurrentIndex(0)
                                    self.ResetAddTab()
                                except:
                                    pass
                            except:
                                QMessageBox.warning(self, "Error", "يرجي التحقق من مساعدات وشهريات الكنائس !!")
                        except:
                            QMessageBox.warning(self, "Error", "يرجي التحقق من صفحة (ملاحظات اخري) !!")
                    except:
                        QMessageBox.warning(self, "Error", "يرجي التحقق من المعلومات !!")
                except:
                    QMessageBox.warning(self, "Error", "يرجي كتابه مساعدات وشهريات الكنائس بدون احرف !!")
        except:
            QMessageBox.warning(self, "Error", "يرجي التحقق من البيانات !!")

    # ------------------------------------------------------------- #
    #                    -- Add New Children --                     #

    def AddChildren(self):
        global d
        NumberFamily = self.lineEdit_16.text()
        ChildrenName = self.lineEdit_26.text()
        ChildrenNationalId = self.lineEdit_25.text()
        ChildrenSocialStatus = self.comboBox_4.currentText()
        ChildrenSchoolWork = self.lineEdit_29.text()
        ChildrenMonthlyIncome = self.lineEdit_28.text()
        ChildrenFatherOfConfession = self.comboBox_15.currentText()
        if NumberFamily == "":
            self.tabWidget_2.setCurrentIndex(0)
            QMessageBox.warning(self, "Error", "(رقم الأسرة) غير موجود في الصفحة الاولي  ...")
        elif ChildrenName == "":
            QMessageBox.warning(self, "Error", "(اسم الابن) غير موجود ...")
        elif ChildrenNationalId == "":
            QMessageBox.warning(self, "Error", "(الرقم القومي) غير موجود  ...")
        elif ChildrenSchoolWork == "":
            QMessageBox.warning(self, "Error", "(السنة الدراسة اول الفصل) غير موجود ...")
        elif ChildrenMonthlyIncome == "":
            QMessageBox.warning(self, "Error", "(الدخل الشهري) غير موجود ...")
        else:
            try:
                if int(len(ChildrenNationalId)) == 14:
                    try:
                        ChildrenMonthlyIncome = int(ChildrenMonthlyIncome)
                        try:
                            NumberFamily = int(NumberFamily)
                            try:
                                self.cur.execute(
                                    f"""SELECT ChildrenNumber FROM anbaabram.children WHERE NumberFamily = "{NumberFamily}" """)
                                Data = self.cur.fetchall()
                                if Data == ():
                                    d = 1
                                for data in Data:
                                    d = int(data[0]) + 1
                                self.cur.execute(
                                    f"""INSERT INTO anbaabram.children (ChildrenName, ChildrenNationalId, ChildrenSocialStatus, ChildrenSchoolWork, ChildrenMonthlyIncome, ChildrenFatherOfConfession, NumberFamily, ChildrenNumber) 
                                            VALUES ("{ChildrenName}", "{ChildrenNationalId}", "{ChildrenSocialStatus}", "{ChildrenSchoolWork}", "{ChildrenMonthlyIncome}", "{ChildrenFatherOfConfession}", "{NumberFamily}", "{d}") """)
                                self.db.commit()
                                self.ResetAddChildren()
                                self.UpdateNumberChildren()
                                self.statusbar.showMessage("تم اضافة بيانات الابن بأسم -- {} --".format(ChildrenName))
                                self.worker = ClearStatusBar()
                                self.worker.start()
                                self.worker.ClearStatusBarVariable.connect(self.StatusBar)

                            except:
                                QMessageBox.warning(self, "Error", "الرقم القومي موجود بالفعل !!")
                        except:
                            QMessageBox.warning(self, "Error", "تأكد ان رقم الاسرة مكون من ارقام فقط !")
                    except:
                        QMessageBox.warning(self, "Error", "تأكد ان الدخل الشهري مكتوب بدون أحرف او علامة عشرية ")
                else:
                    QMessageBox.warning(self, "Error", "الرقم القومي ليس مكون من 14 رقم")
            except:
                QMessageBox.warning(self, "Error", "يجب ان يكون الرقم القومي مكون من ارقام فقط !")

    # ------------------------------------------------------------- #
    #                    -- Add New Human --                        #

    def AddHumanData(self):
        global d
        NumberFamily = self.lineEdit_16.text()
        HumanName = self.lineEdit_32.text()
        HumanNationalId = self.lineEdit_31.text()
        HumanRelativeRelation = self.lineEdit_35.text()
        HumanSocialStatus = self.comboBox_5.currentText()
        HumanSchoolWork = self.lineEdit_34.text()
        HumanMonthlyIncome = self.lineEdit_30.text()
        HumanFatherOfConfession = self.comboBox_22.currentText()
        if NumberFamily == "":
            self.tabWidget_2.setCurrentIndex(0)
            QMessageBox.warning(self, "Error", "(رقم الأسرة) غير موجود في الصفحة الاولي  ...")
        elif HumanName == "":
            QMessageBox.warning(self, "Error", "ادخل (اسم الشخص) ...")
        elif HumanNationalId == "":
            QMessageBox.warning(self, "Error", "ادخل (الرقم القومي) ...")
        elif HumanRelativeRelation == "":
            QMessageBox.warning(self, "Error", "ادخل (صلة القرابة) ...")
        elif HumanSchoolWork == "":
            QMessageBox.warning(self, "Error", "ادخل (السنة الدراسية او الوظيفة) ...")
        elif HumanMonthlyIncome == "":
            QMessageBox.warning(self, "Error", "ادخل (الدخل الشهري) ...")
        else:
            try:
                if int(len(HumanNationalId)) == 14:
                    try:
                        HumanMonthlyIncome = int(HumanMonthlyIncome)
                        try:
                            NumberFamily = int(NumberFamily)
                            try:
                                self.cur.execute(
                                    f"""SELECT HumanNum FROM anbaabram.anotherhuman WHERE NumberFamily = "{NumberFamily}" """)
                                Data = self.cur.fetchall()
                                if Data == ():
                                    d = 1
                                for data in Data:
                                    d = int(data[0]) + 1
                                self.cur.execute(
                                    f"""INSERT INTO anbaabram.anotherhuman (HumanName, HumanNationalId, HumanRelativeRelation, HumanSocialStatus, HumanSchoolWork, HumanMonthlyIncome, HumanFatherOfConfession, NumberFamily, HumanNum) 
                                            VALUES ("{HumanName}", "{HumanNationalId}", "{HumanRelativeRelation}", "{HumanSocialStatus}", "{HumanSchoolWork}", "{HumanMonthlyIncome}", "{HumanFatherOfConfession}", "{NumberFamily}", "{d}") """)
                                self.db.commit()
                                self.ResetAddHuman()
                                self.UpdateNumberOtherHuman()
                                self.statusbar.showMessage("تم اضافة بيانات الشخص بأسم -- {} --".format(HumanName))
                                self.worker = ClearStatusBar()
                                self.worker.start()
                                self.worker.ClearStatusBarVariable.connect(self.StatusBar)

                            except:
                                QMessageBox.warning(self, "Error", "الرقم القومي موجود بالفعل !!")
                        except:
                            QMessageBox.warning(self, "Error", "تأكد ان رقم الاسرة مكون من ارقام فقط !")
                    except:
                        QMessageBox.warning(self, "Error", "تأكد ان الدخل الشهري مكتوب بدون أحرف او علامة عشرية ")
                else:
                    QMessageBox.warning(self, "Error", "الرقم القومي ليس مكون من 14 رقم")
            except:
                QMessageBox.warning(self, "Error", "يجب ان يكون الرقم القومي مكون من ارقام فقط !")

    # ------------------------------------------------------------- #
    #          -- Check Husband and Wife for next page --           #

    def Page1HusbandAndWife(self):
        # -- Husband Data -- # ----------------------------------------------
        HusbandName = self.lineEdit_2.text()
        HusbandNikeName = self.lineEdit_3.text()
        HusbandId = self.lineEdit_4.text()
        HusbandWork = self.lineEdit_5.text()
        HusbandMoney = self.lineEdit_6.text()
        HusbandPhone = self.lineEdit_7.text()
        HusbandState = True
        # -- Wife Data -- # -------------------------------------------------
        WifeName = self.lineEdit_12.text()
        WifeNikeName = self.lineEdit_13.text()
        WifeId = self.lineEdit_9.text()
        WifeWork = self.lineEdit_14.text()
        WifeMoney = self.lineEdit_15.text()
        WifePhone = self.lineEdit_11.text()
        WifeState = True
        if HusbandName != "" or HusbandNikeName != "" or HusbandId != "" or HusbandWork != "" or HusbandMoney != "" or HusbandPhone != "" or WifeName != "" or WifeNikeName != "" or WifeId != "" or WifeWork != "" or WifeMoney != "" or WifePhone != "":
            if HusbandName != "" or HusbandNikeName != "" or HusbandId != "" or HusbandWork != "" or HusbandMoney != "" or HusbandPhone != "":
                if HusbandName != "" and HusbandNikeName != "" and HusbandId != "" and HusbandWork != "" and HusbandMoney != "" and HusbandPhone != "":
                    try:
                        if type(int(HusbandId)) == type(0):
                            if int(len(HusbandId)) == 14:
                                try:
                                    if type(int(HusbandMoney)) == type(0):
                                        try:
                                            if type(int(HusbandPhone)) == type(0):
                                                if int(len(HusbandPhone)) == 11:
                                                    try:
                                                        if self.lineEdit_16.text() != "" and type(
                                                                int(self.lineEdit_16.text())) == type(0):
                                                            HusbandState = True
                                                        else:
                                                            HusbandState = False
                                                            QMessageBox.warning(self, "Error",
                                                                                "تحقق من رقم الأسرة اولا ...!")
                                                    except:
                                                        HusbandState = False
                                                else:
                                                    HusbandState = False
                                                    QMessageBox.warning(self, "Error",
                                                                        "يجب ان يكون رقم الهاتف للزوج مكون من 11 رقم ...!")
                                        except:
                                            HusbandState = False
                                            QMessageBox.warning(self, "Error",
                                                                "يجب كتابة رقم الهاتف للزوج بالارقام فقط ...!")
                                except:
                                    HusbandState = False
                                    QMessageBox.warning(self, "Error",
                                                        "يجب كتابة مرتب الزوج بالارقام فقط بدون حروف او علامة عشرية ...!")
                            else:
                                HusbandState = False
                                QMessageBox.warning(self, "Error", "يجب ان يكون الرقم القومي للزوج مكون من 14 ...!")
                    except:
                        HusbandState = False
                        QMessageBox.warning(self, "Error", "يجب كتابة الرقم القومي للزوج بالارقام فقط ...!")
                else:
                    HusbandState = False
                    QMessageBox.warning(self, "Error", "يرجي اكمال بيانات الزوج ... ")

            if WifeName != "" or WifeNikeName != "" or WifeId != "" or WifeWork != "" or WifeMoney != "" or WifePhone != "":
                if WifeName != "" and WifeNikeName != "" and WifeId != "" and WifeWork != "" and WifeMoney != "" and WifePhone != "":
                    try:
                        if type(int(WifeId)) == type(0):
                            if int(len(WifeId)) == 14:
                                try:
                                    if type(int(WifeMoney)) == type(0):
                                        try:
                                            if type(int(WifePhone)) == type(0):
                                                if int(len(WifePhone)) == 11:
                                                    try:
                                                        if self.lineEdit_16.text() != "" and type(
                                                                int(self.lineEdit_16.text())) == type(0):
                                                            WifeState = True
                                                        else:
                                                            HusbandState = False
                                                            QMessageBox.warning(self, "Error",
                                                                                "تحقق من رقم الأسرة اولا ...!")

                                                    except:
                                                        QMessageBox.warning(self, "Error",
                                                                            "يجب كتابة رقم الأسرة بالارقام فقط ...!")
                                                        WifeState = False
                                                else:
                                                    WifeState = False
                                                    QMessageBox.warning(self, "Error",
                                                                        "يجب ان يكون رقم الهاتف للزوجة مكون من 11 رقم ...!")
                                        except:
                                            WifeState = False
                                            QMessageBox.warning(self, "Error",
                                                                "يجب كتابة رقم الهاتف للزوجة بالارقام فقط ...!")
                                except:
                                    WifeState = False
                                    QMessageBox.warning(self, "Error",
                                                        "يجب كتابة مرتب الزوجة بالارقام فقط بدون حروف او علامة عشرية ...!")
                            else:
                                WifeState = False
                                QMessageBox.warning(self, "Error", "يجب ان يكون الرقم القومي للزوجة مكون من 14 ...!")
                    except:
                        WifeState = False
                        QMessageBox.warning(self, "Error", "يجب كتابة الرقم القومي للزوجة بالارقام فقط ...!")
                else:
                    WifeState = False
                    QMessageBox.warning(self, "Error", "يرجي اكمال بيانات الزوجة ... ")
            if WifeState and HusbandState:
                self.tabWidget_2.setCurrentIndex(1)
                self.ClearLinesFunction()
        else:
            self.tabWidget_2.setCurrentIndex(1)

    # ------------------------------------------------------------- #
    #              -- Clear All Data and reset all --               #

    def ResetAddTab(self):
        self.lineEdit_16.clear()
        self.lineEdit_2.clear()
        self.lineEdit_3.clear()
        self.lineEdit_4.clear()
        self.lineEdit_5.clear()
        self.lineEdit_6.clear()
        self.lineEdit_7.clear()
        self.comboBox_20.setCurrentIndex(0)
        self.lineEdit_12.clear()
        self.lineEdit_13.clear()
        self.lineEdit_9.clear()
        self.lineEdit_14.clear()
        self.lineEdit_15.clear()
        self.lineEdit_11.clear()
        self.textEdit.clear()
        self.comboBox.setCurrentIndex(0)
        self.comboBox_2.setCurrentIndex(0)
        self.lineEdit_19.clear()
        self.lineEdit_20.clear()
        self.radioButton.setChecked(False)
        self.radioButton_2.setChecked(False)
        self.radioButton_4.setChecked(False)
        self.radioButton_3.setChecked(False)
        self.radioButton_5.setChecked(False)
        self.radioButton_6.setChecked(False)
        self.lineEdit_21.setText("كنيسة ")
        self.comboBox_3.setCurrentIndex(0)
        self.lineEdit_24.clear()
        self.lineEdit_23.clear()
        self.lineEdit_26.clear()
        self.lineEdit_25.clear()
        self.comboBox_4.setCurrentIndex(0)
        self.lineEdit_29.clear()
        self.lineEdit_28.clear()
        self.comboBox_15.setCurrentIndex(0)
        self.label_16.setText("0")
        self.lineEdit_32.clear()
        self.lineEdit_31.clear()
        self.lineEdit_35.clear()
        self.comboBox_5.setCurrentIndex(0)
        self.label_22.setText("0")
        self.lineEdit_34.clear()
        self.lineEdit_30.clear()
        self.comboBox_22.setCurrentIndex(0)
        self.textEdit_2.clear()
        self.textEdit_3.clear()
        self.textEdit_4.clear()
        self.comboBox_6.setCurrentIndex(0)
        self.lineEdit_37.clear()
        self.checkBox_3.setChecked(False)
        self.checkBox_5.setChecked(False)
        self.checkBox_4.setChecked(False)
        self.checkBox_6.setChecked(False)
        self.textEdit_5.clear()

    # ------------------------------------------------------------- #
    #              -- Clear And Reset Human Page  --                #

    def ResetAddHuman(self):
        self.lineEdit_32.clear()
        self.lineEdit_31.clear()
        self.lineEdit_35.clear()
        self.comboBox_5.setCurrentIndex(0)
        self.lineEdit_34.clear()
        self.lineEdit_30.clear()
        self.comboBox_22.setCurrentIndex(0)

    # ------------------------------------------------------------- #
    #              -- Clear And Reset Children Page  --             #

    def ResetAddChildren(self):
        self.lineEdit_26.clear()
        self.lineEdit_25.clear()
        self.comboBox_4.setCurrentIndex(0)
        self.lineEdit_29.clear()
        self.lineEdit_28.clear()
        self.comboBox_15.setCurrentIndex(0)

    # ----------------------------------------------------------------- #
    # ----------------------------------------------------------------- #
    #                                                                   #
    #                            -- Edit Tab --                         #
    #                                                                   #
    # ----------------------------------------------------------------- #
    # ----------------------------------------------------------------- #

    # ------------------------------------------------------------- #
    #       -- Get All Data (Search for all data was save)  --      #

    def SearchEdit(self):
        self.ResetEditTab(1)
        global FamilyNum, S
        combo_search = self.comboBox_13.currentIndex()
        Search = self.lineEdit_57.text()
        if Search != "":
            try:
                int(Search)
                if combo_search == 0:
                    FamilyNum = Search
                    S1 = True
                try:
                    if combo_search == 1:
                        self.cur.execute(
                            f"""SELECT NumberFamily FROM anbaabram.husbanddata WHERE HusbandNationalId = "{Search}" """)
                        FamilyNum = self.cur.fetchone()[0]
                        S2 = True
                except:
                    S2 = False
                try:
                    if combo_search == 2:
                        self.cur.execute(
                            f"""SELECT NumberFamily FROM anbaabram.wifedata WHERE WifeNationalId = "{Search}" """)
                        FamilyNum = self.cur.fetchone()[0]
                        S3 = True
                except:
                    S3 = False
                if S1 or S2 or S3:
                    try:
                        self.cur.execute(
                            f"""SELECT HusbandName, HusbandNikeName, HusbandNationalId, HusbandWork, HusbandMoney, HusbandPhoneNumber, HusbandFatherOfConfession_id FROM anbaabram.husbanddata WHERE NumberFamily = "{FamilyNum}" """)
                        HusbandData = self.cur.fetchone()
                        self.lineEdit_49.setText(HusbandData[0])
                        self.lineEdit_50.setText(HusbandData[1])
                        self.lineEdit_42.setText(HusbandData[2])
                        self.lineEdit_53.setText(HusbandData[3])
                        self.lineEdit_54.setText(HusbandData[4])
                        self.lineEdit_46.setText(HusbandData[5])
                        try:
                            self.cur.execute(
                                f"""INSERT INTO anbaabram.fatherofconfession (FatherOfConfession) VALUES ("{HusbandData[6]}") """)
                            self.db.commit()
                            self.comboBox_17.addItem(HusbandData[6])
                        except:
                            pass
                        self.comboBox_17.setCurrentText(HusbandData[6])
                    except:
                        pass
                    # --------------------------------------------------- #
                    try:
                        self.cur.execute(
                            f"""SELECT WifeName, WifeNikeName, WifeNationalId, WifeWork, WifeMoney, WifePhoneNumber, WifeFatherOfConfession_id FROM anbaabram.wifedata WHERE NumberFamily = "{FamilyNum}" """)
                        WifeData = self.cur.fetchone()
                        self.lineEdit_51.setText(WifeData[0])
                        self.lineEdit_48.setText(WifeData[1])
                        self.lineEdit_44.setText(WifeData[2])
                        self.lineEdit_47.setText(WifeData[3])
                        self.lineEdit_43.setText(WifeData[4])
                        self.lineEdit_56.setText(WifeData[5])
                        try:
                            self.cur.execute(
                                f"""INSERT INTO anbaabram.fatherofconfession (FatherOfConfession) VALUES ("{WifeData[6]}") """)
                            self.db.commit()
                            self.comboBox_18.addItem(WifeData[6])
                        except:
                            pass
                        self.comboBox_18.setCurrentText(WifeData[6])
                    except:
                        pass
                    self.cur.execute(
                        f"""SELECT DetailedAddress, Governorate, neighborhood, Region, SpecialMarque, Living, ChurchName, Address, Phone1, Phone2 FROM anbaabram.information WHERE NumFamily = "{Search}" """)
                    InfoData = self.cur.fetchone()
                    self.textEdit_6.setText(InfoData[0])
                    try:
                        self.cur.execute(
                            f"""INSERT INTO anbaabram.governorate (GovernorateName) VALUES ("{InfoData[1]}") """)
                        self.db.commit()
                        self.comboBox_9.addItem(InfoData[1])
                    except:
                        pass
                    self.comboBox_9.setCurrentText(InfoData[1])
                    try:
                        self.cur.execute(f"""INSERT INTO anbaabram.district (DistrictName) VALUES ("{InfoData[2]}") """)
                        self.db.commit()
                        self.comboBox_8.addItem(InfoData[2])
                    except:
                        pass
                    self.comboBox_8.setCurrentText(InfoData[2])

                    self.lineEdit_106.setText(InfoData[3])
                    self.lineEdit_103.setText(InfoData[4])

                    if InfoData[5] == self.radioButton_18.text():
                        self.radioButton_18.setChecked(True)

                    elif InfoData[5] == self.radioButton_17.text():
                        self.radioButton_17.setChecked(True)

                    elif InfoData[5] == self.radioButton_19.text():
                        self.radioButton_19.setChecked(True)

                    elif InfoData[5] == self.radioButton_20.text():
                        self.radioButton_20.setChecked(True)
                    if InfoData[6] == self.radioButton_7.text():
                        self.radioButton_7.setChecked(True)
                    else:
                        self.radioButton_8.setChecked(True)
                        self.lineEdit_101.setText(InfoData[6])
                    try:
                        self.cur.execute(f"""INSERT INTO anbaabram.address (addressname) VALUES ("{InfoData[7]}") """)
                        self.db.commit()
                        self.comboBox_7.addItem(InfoData[7])
                    except:
                        pass
                    self.comboBox_7.setCurrentText(InfoData[7])
                    self.lineEdit_105.setText(InfoData[8])
                    self.lineEdit_102.setText(InfoData[9])
                    self.ShowEditHomeContents()
                    self.cur.execute(
                        f"""SELECT Diseases, Obstruction, OtherCircumstances FROM anbaabram.anotherinfofamily WHERE NumFamily = "{FamilyNum}" """)
                    OtherInfo = self.cur.fetchone()
                    self.textEdit_7.setText(OtherInfo[0])
                    self.textEdit_8.setText(OtherInfo[1])
                    self.textEdit_9.setText(OtherInfo[2])
                    self.cur.execute(
                        f"""SELECT SourceOfIncome, MonthlyIncome FROM anbaabram.primarysourcesofincome WHERE NumFamily = "{FamilyNum}" """)
                    PrimarySourcesOfIncome = self.cur.fetchone()
                    try:
                        self.cur.execute(
                            f"""INSERT INTO anbaabram.source (SourceName) VALUES ("{PrimarySourcesOfIncome[0]}") """)
                        self.db.commit()
                        self.comboBox_12.addItem(PrimarySourcesOfIncome[0])
                    except:
                        pass
                    self.comboBox_12.addItem(PrimarySourcesOfIncome[0])
                    self.lineEdit_78.setText(PrimarySourcesOfIncome[1])
                    self.cur.execute(
                        f"""SELECT MonthlyMaterialAssistance, CurativeHelp, MonthlyGood, OtherAid, Others FROM anbaabram.churchesaidandgoods WHERE NumFamily = "{FamilyNum}" """)
                    ChurchEsaidAndGoods = self.cur.fetchone()
                    if ChurchEsaidAndGoods[0] == "1":
                        self.checkBox_9.setChecked(True)
                    else:
                        self.checkBox_9.setChecked(False)
                    if ChurchEsaidAndGoods[1] == "1":
                        self.checkBox_11.setChecked(True)
                    else:
                        self.checkBox_11.setChecked(False)
                    if ChurchEsaidAndGoods[2] == "1":
                        self.checkBox_10.setChecked(True)
                    else:
                        self.checkBox_10.setChecked(False)
                    if ChurchEsaidAndGoods[3] == "1":

                        self.checkBox_12.setChecked(True)
                    else:
                        self.checkBox_12.setChecked(False)

                    self.textEdit_10.setText(ChurchEsaidAndGoods[4])
                    if self.lineEdit_49.text() == "":
                        Status = False
                    else:
                        Status = True
                    self.lineEdit_49.setEnabled(Status)
                    self.lineEdit_50.setEnabled(Status)
                    self.lineEdit_42.setEnabled(Status)
                    self.lineEdit_53.setEnabled(Status)
                    self.lineEdit_54.setEnabled(Status)
                    self.lineEdit_46.setEnabled(Status)
                    self.comboBox_17.setEnabled(Status)
                    if self.lineEdit_51.text() == "":
                        Status = False
                    else:
                        Status = True
                    self.lineEdit_51.setEnabled(Status)
                    self.lineEdit_48.setEnabled(Status)
                    self.lineEdit_44.setEnabled(Status)
                    self.lineEdit_47.setEnabled(Status)
                    self.lineEdit_43.setEnabled(Status)
                    self.lineEdit_56.setEnabled(Status)
                    self.comboBox_18.setEnabled(Status)

                    try:
                        self.cur.execute(
                            f"""SELECT * FROM anbaabram.churchesaidandgoodspart2 WHERE NumFamily = "{FamilyNum}" """)
                        ChurchEsaidAndGoodspart2 = self.cur.fetchone()
                        ChurchEsaidAndGoodspart2 = list(ChurchEsaidAndGoodspart2)
                        self.lineEdit_88.setText(ChurchEsaidAndGoodspart2[1])
                        self.lineEdit_87.setText(ChurchEsaidAndGoodspart2[2])
                        self.lineEdit_86.setText(ChurchEsaidAndGoodspart2[3])
                        self.lineEdit_94.setText(ChurchEsaidAndGoodspart2[4])
                        self.lineEdit_89.setText(ChurchEsaidAndGoodspart2[5])
                        self.lineEdit_90.setText(ChurchEsaidAndGoodspart2[7])
                        self.lineEdit_85.setText(ChurchEsaidAndGoodspart2[8])
                        self.lineEdit_92.setText(ChurchEsaidAndGoodspart2[9])
                        self.lineEdit_91.setText(ChurchEsaidAndGoodspart2[10])
                        self.lineEdit_93.setText(ChurchEsaidAndGoodspart2[11])
                    except:
                        pass


            except:
                self.EnabledLinesH_W()
                QMessageBox.warning(self, "Error", "الرقم الذي ادخلته غير موجود ...!")
        else:
            self.EnabledLinesH_W()
            QMessageBox.warning(self, "Error", "يرجي ادخال رقم الاسرة ...!")

    def EnabledLinesH_W(self):
        Status = True
        self.lineEdit_49.setEnabled(Status)
        self.lineEdit_50.setEnabled(Status)
        self.lineEdit_42.setEnabled(Status)
        self.lineEdit_53.setEnabled(Status)
        self.lineEdit_54.setEnabled(Status)
        self.lineEdit_46.setEnabled(Status)
        self.comboBox_17.setEnabled(Status)
        self.lineEdit_51.setEnabled(Status)
        self.lineEdit_48.setEnabled(Status)
        self.lineEdit_44.setEnabled(Status)
        self.lineEdit_47.setEnabled(Status)
        self.lineEdit_43.setEnabled(Status)
        self.lineEdit_56.setEnabled(Status)
        self.comboBox_18.setEnabled(Status)

    # ------------------------------------------------------------- #
    #   -- Save Husband and Wife with new data (Update Data)  --    #

    def UpdateHusbandAndWifeData(self):
        # -- number family -- # ----------------------------------------------
        global ChurchName, Living
        NumFamily = self.lineEdit_57.text()
        # -- Husband Data -- # ----------------------------------------------
        HusbandName = self.lineEdit_49.text()
        HusbandNikeName = self.lineEdit_50.text()
        HusbandId = self.lineEdit_42.text()
        HusbandWork = self.lineEdit_53.text()
        HusbandMoney = self.lineEdit_54.text()
        HusbandPhone = self.lineEdit_46.text()
        HusbandFatherOfConfession = self.comboBox_17.currentText()
        HusbandState = True
        # -- Wife Data -- # ---------------------------------------------------
        WifeName = self.lineEdit_51.text()
        WifeNikeName = self.lineEdit_48.text()
        WifeId = self.lineEdit_44.text()
        WifeWork = self.lineEdit_47.text()
        WifeMoney = self.lineEdit_43.text()
        WifePhone = self.lineEdit_56.text()
        WifeFatherOfConfession = self.comboBox_18.currentText()
        WifeState = True
        if HusbandName != "" or HusbandNikeName != "" or HusbandId != "" or HusbandWork != "" or HusbandMoney != "" or HusbandPhone != "" or WifeName != "" or WifeNikeName != "" or WifeId != "" or WifeWork != "" or WifeMoney != "" or WifePhone != "":
            if HusbandName != "" or HusbandNikeName != "" or HusbandId != "" or HusbandWork != "" or HusbandMoney != "" or HusbandPhone != "":
                if HusbandName != "" and HusbandNikeName != "" and HusbandId != "" and HusbandWork != "" and HusbandMoney != "" and HusbandPhone != "":
                    try:
                        if type(int(HusbandId)) == type(0):
                            if int(len(HusbandId)) == 14:
                                try:
                                    if type(int(HusbandMoney)) == type(0):
                                        try:
                                            if type(int(HusbandPhone)) == type(0):
                                                if int(len(HusbandPhone)) == 11:
                                                    try:
                                                        if self.lineEdit_57.text() != "" and type(
                                                                int(self.lineEdit_57.text())) == type(0):
                                                            HusbandState = True
                                                        else:
                                                            HusbandState = False
                                                            QMessageBox.warning(self, "Error",
                                                                                "تحقق من رقم الأسرة اولا ...!")
                                                    except:
                                                        HusbandState = False
                                                else:
                                                    HusbandState = False
                                                    QMessageBox.warning(self, "Error",
                                                                        "يجب ان يكون رقم الهاتف للزوج مكون من 11 رقم ...!")
                                        except:
                                            HusbandState = False
                                            QMessageBox.warning(self, "Error",
                                                                "يجب كتابة رقم الهاتف للزوج بالارقام فقط ...!")
                                except:
                                    HusbandState = False
                                    QMessageBox.warning(self, "Error",
                                                        "يجب كتابة مرتب الزوج بالارقام فقط بدون حروف او علامة عشرية ...!")
                            else:
                                HusbandState = False
                                QMessageBox.warning(self, "Error", "يجب ان يكون الرقم القومي للزوج مكون من 14 ...!")
                    except:
                        HusbandState = False
                        QMessageBox.warning(self, "Error", "يجب كتابة الرقم القومي للزوج بالارقام فقط ...!")
                else:
                    HusbandState = False
                    QMessageBox.warning(self, "Error", "يرجي اكمال بيانات الزوج ... ")

            if WifeName != "" or WifeNikeName != "" or WifeId != "" or WifeWork != "" or WifeMoney != "" or WifePhone != "":
                if WifeName != "" and WifeNikeName != "" and WifeId != "" and WifeWork != "" and WifeMoney != "" and WifePhone != "":
                    try:
                        if type(int(WifeId)) == type(0):
                            if int(len(WifeId)) == 14:
                                try:
                                    if type(int(WifeMoney)) == type(0):
                                        try:
                                            if type(int(WifePhone)) == type(0):
                                                if int(len(WifePhone)) == 11:
                                                    try:
                                                        if self.lineEdit_57.text() != "" and type(
                                                                int(self.lineEdit_57.text())) == type(0):
                                                            WifeState = True
                                                        else:
                                                            HusbandState = False
                                                            QMessageBox.warning(self, "Error",
                                                                                "تحقق من رقم الأسرة اولا ...!")

                                                    except:
                                                        QMessageBox.warning(self, "Error",
                                                                            "يجب كتابة رقم الأسرة بالارقام فقط ...!")
                                                        WifeState = False
                                                else:
                                                    WifeState = False
                                                    QMessageBox.warning(self, "Error",
                                                                        "يجب ان يكون رقم الهاتف للزوجة مكون من 11 رقم ...!")
                                        except:
                                            WifeState = False
                                            QMessageBox.warning(self, "Error",
                                                                "يجب كتابة رقم الهاتف للزوجة بالارقام فقط ...!")
                                except:
                                    WifeState = False
                                    QMessageBox.warning(self, "Error",
                                                        "يجب كتابة مرتب الزوجة بالارقام فقط بدون حروف او علامة عشرية ...!")
                            else:
                                WifeState = False
                                QMessageBox.warning(self, "Error", "يجب ان يكون الرقم القومي للزوجة مكون من 14 ...!")
                    except:
                        WifeState = False
                        QMessageBox.warning(self, "Error", "يجب كتابة الرقم القومي للزوجة بالارقام فقط ...!")
                else:
                    WifeState = False
                    QMessageBox.warning(self, "Error", "يرجي اكمال بيانات الزوجة ... ")
            if WifeState and HusbandState:
                try:
                    self.cur.execute(
                        f"""UPDATE anbaabram.husbanddata SET HusbandName = "{HusbandName}", HusbandNikeName = "{HusbandNikeName}", HusbandNationalId = "{HusbandId}", HusbandWork= "{HusbandWork}", HusbandMoney = "{HusbandMoney}", HusbandPhoneNumber = "{HusbandPhone}", HusbandFatherOfConfession_id = "{HusbandFatherOfConfession}" WHERE NumberFamily = "{NumFamily}" """)
                    self.db.commit()
                except:
                    HusbandState = False
                try:
                    self.cur.execute(
                        f"""UPDATE anbaabram.wifedata SET WifeName = "{WifeName}", WifeNikeName = "{WifeNikeName}", WifeNationalId = "{WifeId}", WifeWork = "{WifeWork}", WifeMoney = "{WifeMoney}", WifePhoneNumber = "{WifePhone}", WifeFatherOfConfession_id = "{WifeFatherOfConfession}" WHERE NumberFamily = "{NumFamily}" """)
                    self.db.commit()
                except:
                    WifeState = False
                if HusbandState and WifeState:
                    QMessageBox.information(self, "Done", "تم حفظ البيانات بنجاح")
                else:
                    QMessageBox.warning(self, "Error", "يرجي التأكد من البيانات !!")

    # ------------------------------------------------------------- #
    #         -- Save Info with new data (Update Data)  --          #

    def SaveInfo(self):
        global ChurchName, Living
        NumFamily = self.lineEdit_57.text()
        if NumFamily != "":
            try:
                DetailedAddress = self.textEdit_6.toPlainText()
                Governorate = self.comboBox_9.currentText()
                neighborhood = self.comboBox_8.currentText()
                Region = self.lineEdit_106.text()
                SpecialMarque = self.lineEdit_103.text()
                if self.radioButton_18.isChecked():
                    Living = self.radioButton_18.text()
                elif self.radioButton_17.isChecked():
                    Living = self.radioButton_17.text()
                elif self.radioButton_19.isChecked():
                    Living = self.radioButton_19.text()
                elif self.radioButton_20.isChecked():
                    Living = self.radioButton_20.text()
                else:
                    QMessageBox.warning(self, "Error", "تحقق من نظام السكن ...")
                if self.radioButton_7.isChecked() is True:
                    ChurchName = self.radioButton_7.text()
                elif self.radioButton_8.isChecked() is True:
                    ChurchName = self.lineEdit_101.text()
                else:
                    QMessageBox.warning(self, "Error", "تحقق من اسم الكنيسة ...")
                Address = self.comboBox_7.currentText()
                Phone1 = self.lineEdit_105.text()
                Phone2 = self.lineEdit_102.text()
                Status = True
                # ------------------------------------------------------
                if DetailedAddress == "" or Region == "" or SpecialMarque == "" or Phone1 == "" or Phone2 == "":
                    question = QMessageBox.question(self, "Warning", "هل انت متأكد من حفظ البيانت غير المسجلة ؟",
                                                    QMessageBox.Yes | QMessageBox.No)
                    if question == QMessageBox.Yes:
                        Status = True
                    else:
                        Status = False
                if Status:
                    self.cur.execute(f"""UPDATE anbaabram.information SET 
                        DetailedAddress = "{DetailedAddress}", 
                        Governorate = "{Governorate}", 
                        neighborhood = "{neighborhood}", 
                        Region = "{Region}", 
                        SpecialMarque = "{SpecialMarque}", 
                        Living = "{Living}", 
                        ChurchName = "{ChurchName}", 
                        Address = "{Address}", 
                        Phone1 = "{Phone1}", 
                        Phone2 = "{Phone2}" 
                        WHERE NumFamily = "{NumFamily}" """)
                    self.db.commit()
                    self.ResetEditTab(3)
                    QMessageBox.information(self, "Done", "تم حفظ البيانات بنجاح برقم اسرة ({})".format(NumFamily))


            except:
                QMessageBox.warning(self, "Error", "رقم الاسرة غير مسجل بالفعل ...\n يرجي التحقق منه")

        else:
            self.tabWidget_5.setCurrentIndex(0)
            QMessageBox.warning(self, "Error", "يرجي كتابه رقم الاسرة اولا ...")

    # ------------------------------------------------------------- #
    #                 -- Update Other Info Data  --                 #

    def UpdateOtherInfoData(self):
        NumFamily = self.lineEdit_57.text()
        if NumFamily != "":
            try:
                Diseases = self.textEdit_7.toPlainText()
                Obstruction = self.textEdit_8.toPlainText()
                OtherCircumstances = self.textEdit_9.toPlainText()
                self.cur.execute(f"""UPDATE anbaabram.anotherinfofamily SET 
                    Diseases = "{Diseases}", 
                    Obstruction = "{Obstruction}", 
                    OtherCircumstances = "{OtherCircumstances}"                     
                    WHERE NumFamily = "{NumFamily}" """)
                self.db.commit()
                QMessageBox.information(self, "Done", "تم حفظ البيانات بنجاح برقم اسرة ({})".format(NumFamily))
            except:
                QMessageBox.warning(self, "Error", "رقم الاسرة غير مسجل بالفعل ...\n يرجي التحقق منه")
        else:
            self.tabWidget_5.setCurrentIndex(0)
            QMessageBox.warning(self, "Error", "يرجي كتابه رقم الاسرة اولا ...")

    # ------------------------------------------------------------- #
    #                       -- Save All Data --                     #

    def SaveLastPage(self):
        NumFamily = self.lineEdit_57.text()

        SourceOfIncome = self.comboBox_12.currentText()
        MonthlyIncome = self.lineEdit_78.text()

        MonthlyMaterialAssistance = self.checkBox_9.checkState().__str__()
        CurativeHelp = self.checkBox_11.checkState().__str__()
        MonthlyGood = self.checkBox_10.checkState().__str__()
        OtherAid = self.checkBox_12.checkState().__str__()
        S_H1 = "0"
        S_H3 = "0"
        S_H2 = "0"
        S_H4 = "0"
        if MonthlyMaterialAssistance == "2":
            S_H1 = "1"
        if CurativeHelp == "2":
            S_H2 = "1"
        if MonthlyGood == "2":
            S_H3 = "1"
        if OtherAid == "2":
            S_H4 = "1"
        Others = self.textEdit_10.toPlainText()
        Status = True
        if NumFamily != "":
            if MonthlyIncome == "":
                question = QMessageBox.question(self, "Warning", "هل انت متأكد من حفظ البيانت غير المسجلة ؟",
                                                QMessageBox.Yes | QMessageBox.No)
                if question == QMessageBox.Yes:
                    Status = True
                else:
                    Status = False
            if Status:
                self.cur.execute(f"""UPDATE anbaabram.primarysourcesofincome SET 
                    SourceOfIncome = "{SourceOfIncome}", 
                    MonthlyIncome = "{MonthlyIncome}" 
                    WHERE NumFamily = "{NumFamily}" """)
                self.db.commit()
                self.cur.execute(f"""UPDATE anbaabram.churchesaidandgoods SET 
                    MonthlyMaterialAssistance = "{S_H1}", 
                    CurativeHelp = "{S_H2}", 
                    MonthlyGood = "{S_H3}", 
                    OtherAid = "{S_H4}", 
                    Others = "{Others}" 
                    WHERE NumFamily = "{NumFamily}" """)
                self.db.commit()
                self.ResetEditTab(5)
                QMessageBox.information(self, "Done", "تم حفظ البيانات بنجاح برقم اسرة ({})".format(NumFamily))
        else:
            self.tabWidget_5.setCurrentIndex(0)
            QMessageBox.warning(self, "Error", "يرجي كتابه رقم الاسرة اولا ...")

    # ------------------------------------------------------------- #
    #                       -- Save All Data --                     #

    def SaveLastPagePart2(self):
        NumFamily = self.lineEdit_57.text()

        ChurchesAidAndGoods = self.lineEdit_88.text()
        TherapeuticAid = self.lineEdit_87.text()
        AidDuringTheStudy = self.lineEdit_86.text()
        BasicSalary = self.lineEdit_94.text()
        AdditionalResource_project = self.lineEdit_89.text()

        ElectricityWaterAndGas = self.lineEdit_90.text()
        Phone = self.lineEdit_85.text()
        Rent = self.lineEdit_92.text()
        Therapy = self.lineEdit_91.text()
        AStudy = self.lineEdit_93.text()

        try:
            Total1 = float(ChurchesAidAndGoods) + float(TherapeuticAid) + float(AidDuringTheStudy) + float(
                BasicSalary) + float(AdditionalResource_project)
            Total2 = float(ElectricityWaterAndGas) + float(Phone) + float(Rent) + float(Therapy) + float(AStudy)
            self.cur.execute(f"""UPDATE anbaabram.churchesaidandgoodspart2 SET 
            ChurchesAidAndGoods = "{ChurchesAidAndGoods}", 
            TherapeuticAid = "{TherapeuticAid}", 
            AidDuringTheStudy = "{AidDuringTheStudy}", 
            BasicSalary = "{BasicSalary}", 
            AdditionalResource_project = "{AdditionalResource_project}", 
            Total1 = "{Total1}", 
            ElectricityWaterAndGas = "{ElectricityWaterAndGas}", 
            Phone = "{Phone}", 
            Rent = "{Rent}", 
            Therapy = "{Therapy}", 
            AStudy = "{AStudy}", 
            Total2 = "{Total2}" 
            WHERE NumFamily = "{NumFamily}" """)
            self.db.commit()
            self.statusbar.showMessage("تم تحديث البيانات بنجاح")
            self.worker = ClearStatusBar()
            self.worker.start()
            self.worker.ClearStatusBarVariable.connect(self.StatusBar)
        except:
            QMessageBox.warning(self, "Error", "يرجي كتابه مساعدات وشهريات الكنائس بدون احرف !!")

    # ------------------------------------------------------------- #
    #                      -- Delete All Data --                    #

    def DeleteAllData(self):
        NumFamily = self.lineEdit_57.text()
        if NumFamily != "":
            question = QMessageBox.question(self, "Sure", "هل انت متأكد من مسح كل البيانات ؟",
                                            QMessageBox.Yes | QMessageBox.No)
            if question == QMessageBox.Yes:
                Status = True
            else:
                Status = False
            if Status:
                self.P_B = ProgressBar()
                self.progressBar.setVisible(True)
                self.Show_ProgressBar_Delete()
                self.P_B.start()
                self.P_B.prog.connect(self.Update_progressBar)
                try:
                    self.cur.execute(f"""DELETE FROM anbaabram.husbanddata WHERE NumberFamily = "{NumFamily}" """)
                    self.db.commit()
                except:
                    pass
                try:
                    self.cur.execute(f"""DELETE FROM anbaabram.wifedata WHERE NumberFamily = "{NumFamily}" """)
                    self.db.commit()
                except:
                    pass
                try:
                    self.cur.execute(f"""DELETE FROM anbaabram.information WHERE NumFamily = "{NumFamily}" """)
                    self.db.commit()
                except:
                    pass
                try:
                    self.cur.execute(f"""DELETE FROM anbaabram.children WHERE NumberFamily = "{NumFamily}" """)
                    self.db.commit()
                except:
                    pass
                try:
                    self.cur.execute(f"""DELETE FROM anbaabram.anotherhuman WHERE NumberFamily = "{NumFamily}" """)
                    self.db.commit()
                except:
                    pass
                try:
                    self.cur.execute(f"""DELETE FROM anbaabram.hcontents WHERE FamilyNum = "{NumFamily}" """)
                    self.db.commit()
                except:
                    pass
                try:
                    self.cur.execute(f"""DELETE FROM anbaabram.anotherinfofamily WHERE NumFamily = "{NumFamily}" """)
                    self.db.commit()
                except:
                    pass
                try:
                    self.cur.execute(
                        f"""DELETE FROM anbaabram.primarysourcesofincome WHERE NumFamily = "{NumFamily}" """)
                    self.db.commit()
                except:
                    pass
                try:
                    self.cur.execute(f"""DELETE FROM anbaabram.churchesaidandgoods WHERE NumFamily = "{NumFamily}" """)
                    self.db.commit()
                except:
                    pass
                try:
                    self.cur.execute(
                        f"""DELETE FROM anbaabram.churchesaidandgoodspart2 WHERE NumFamily = "{NumFamily}" """)
                    self.db.commit()
                except:
                    pass

                self.lineEdit_57.clear()
                self.ResetEditTab(1)
                QMessageBox.information(self, "Done", "تم مسح البيانات بنجاح")
        else:
            QMessageBox.warning(self, "Error", "يرجي كتابه رقم الاسرة اولا ...")

    # ------------------------------------------------------------- #
    def Show_ProgressBar_Delete(self):
        p = QPropertyAnimation(self.progressBar, b"geometry")
        p.setDuration(500)
        p.setStartValue(QRect(840, 70, 140, 20))
        p.setEndValue(QRect(840, 90, 140, 20))
        p.start()
        self.p = p

    def High_ProgressBar_Delete(self):
        p = QPropertyAnimation(self.progressBar, b"geometry")
        p.setDuration(500)
        p.setStartValue(QRect(840, 90, 140, 20))
        p.setEndValue(QRect(840, 70, 140, 20))
        p.start()
        self.p = p

    def Update_progressBar(self, val):
        self.progressBar.setValue(val)
        if val >= 70:
            self.High_ProgressBar_Delete()

    # ------------------------------------------------------------- #

    # ------------------------------------------------------------- #
    #               -- Clear All Data And Reset All --              #

    def ResetEditTab(self, Range):
        if Range == 1 or 2:
            self.lineEdit_49.clear()
            self.lineEdit_50.clear()
            self.lineEdit_42.clear()
            self.lineEdit_53.clear()
            self.lineEdit_54.clear()
            self.lineEdit_46.clear()
            self.comboBox_17.setCurrentIndex(0)
            self.lineEdit_51.clear()
            self.lineEdit_48.clear()
            self.lineEdit_44.clear()
            self.lineEdit_47.clear()
            self.lineEdit_43.clear()
            self.lineEdit_56.clear()
            self.comboBox_18.setCurrentIndex(0)
        if Range == 1 or 3:
            self.textEdit_6.clear()
            self.comboBox_9.setCurrentIndex(0)
            self.comboBox_8.setCurrentIndex(0)
            self.lineEdit_106.clear()
            self.lineEdit_103.clear()
            self.radioButton_18.setChecked(False)
            self.radioButton_17.setChecked(False)
            self.radioButton_19.setChecked(False)
            self.radioButton_20.setChecked(False)
            self.radioButton_7.setChecked(False)
            self.radioButton_8.setChecked(False)
            self.lineEdit_101.setText("كنيسة ")
            self.comboBox_7.setCurrentIndex(0)
            self.lineEdit_105.clear()
            self.lineEdit_102.clear()
        if Range == 1 or 4:
            self.textEdit_7.clear()
            self.textEdit_8.clear()
            self.textEdit_9.clear()
        if Range == 1 or 5:
            self.comboBox_12.setCurrentIndex(0)
            self.lineEdit_78.clear()
            self.checkBox_9.setChecked(False)
            self.checkBox_11.setChecked(False)
            self.checkBox_10.setChecked(False)
            self.checkBox_12.setChecked(False)
            self.textEdit_10.clear()

    # ----------------------------------------------------------------- #
    #                           -- @ Children --                        #
    # ----------------------------------------------------------------- #
    # ------------------------------------------------------------- #
    #                   -- Found All Children --                    #

    def SearchChildren(self):
        global ChildrenNumber
        ChildIdSearch = self.lineEdit_109.text()
        if ChildIdSearch != "":
            try:
                self.cur.execute(
                    f"""SELECT ChildrenName, ChildrenNationalId, ChildrenSocialStatus, ChildrenSchoolWork, ChildrenMonthlyIncome, ChildrenFatherOfConfession FROM anbaabram.children WHERE ChildrenNationalId = "{ChildIdSearch}" """)
                ChildData = self.cur.fetchone()
                self.lineEdit_68.setText(ChildData[0])
                self.lineEdit_70.setText(ChildData[1])
                self.comboBox_10.setCurrentText(ChildData[2])
                self.lineEdit_66.setText(ChildData[3])
                self.lineEdit_67.setText(ChildData[4])
                try:
                    self.cur.execute(
                        f"""INSERT INTO anbaabram.fatherofconfession (FatherOfConfession) VALUES ("{ChildData[5]}") """)
                    self.db.commit()
                    self.comboBox_16.addItem(ChildData[5])
                except:
                    pass
                self.comboBox_16.setCurrentText(ChildData[5])
                self.cur.execute(
                    f"""SELECT NumberFamily FROM anbaabram.children WHERE ChildrenNationalId = "{ChildIdSearch}" """)
                NumFamily = self.cur.fetchone()[0]
                self.cur.execute(
                    f"""SELECT ChildrenNumber FROM anbaabram.children WHERE NumberFamily = "{NumFamily}" """)
                Numbers = self.cur.fetchall()
                for Ch_N in Numbers:
                    ChildrenNumber = Ch_N[0]
                self.lcdNumber_5.setProperty("intValue", int(ChildrenNumber))
            except:
                QMessageBox.warning(self, "Error", "الرقم القومي غير صحيح !!")
        else:
            self.ResetEditChildren()

    # ------------------------------------------------------------- #
    #               -- Save Children With New Data  --              #

    def UpdateChildren(self):
        OldId = self.lineEdit_109.text()
        ChildName = self.lineEdit_68.text()
        ChildId = self.lineEdit_70.text()
        ChildStatusCom = self.comboBox_10.currentText()
        ChildSchool = self.lineEdit_66.text()
        ChildWork = self.lineEdit_67.text()
        ChildFatherOfConfession = self.comboBox_16.currentText()
        if OldId != "" and ChildName != "" and ChildId != "" and ChildSchool != "" and ChildWork != "":
            try:
                int(ChildId)
                try:
                    int(ChildWork)
                    if int(len(ChildId)) == 14:
                        try:
                            self.cur.execute(f"""UPDATE anbaabram.children SET 
                                ChildrenName = "{ChildName}", 
                                ChildrenNationalId = "{ChildId}", 
                                ChildrenSocialStatus = "{ChildStatusCom}", 
                                ChildrenSchoolWork = "{ChildSchool}", 
                                ChildrenMonthlyIncome = "{ChildWork}", 
                                ChildrenFatherOfConfession = "{ChildFatherOfConfession}" 

                                WHERE ChildrenNationalId = "{OldId}" """)
                            self.db.commit()
                            self.statusbar.showMessage("تم تحديث البيانات بنجاح بأسم -- {} --".format(ChildName))
                            self.worker.start()
                            self.worker.ClearStatusBarVariable.connect(self.StatusBar)

                            self.ResetEditChildren()
                        except:
                            QMessageBox.warning(self, "Error", "الرقم القومي الاساسي غير صحيح ...!")
                    else:
                        QMessageBox.warning(self, "Error", "يجب ان يكون الرقم القومي مكون من 14 رقم ...!")
                except:
                    QMessageBox.warning(self, "Error", "الدخل الشهري ليس مكون من ارقام فقط !!")
            except:
                QMessageBox.warning(self, "Error", "الرقم القومي ليس مكون من ارقام فقط !!")
        else:
            QMessageBox.warning(self, "Error", "يرجي اكمال البيانات لحفظها")

    # ------------------------------------------------------------- #
    #                 -- Delete All Children Data --                #

    def DeleteChildrenData(self):
        global row
        ChildrenNationalId = self.lineEdit_109.text()
        if ChildrenNationalId != "":
            try:
                int(ChildrenNationalId)
                self.cur.execute(
                    f"""SELECT * FROM anbaabram.children WHERE ChildrenNationalId = "{ChildrenNationalId}" """)
                if self.cur.fetchall() == ():
                    QMessageBox.warning(self, "Error", "الرقم القومي غير مسجل بالفعل !!")
                else:
                    self.cur.execute(
                        f"""SELECT NumberFamily FROM anbaabram.children WHERE ChildrenNationalId = "{ChildrenNationalId}" """)
                    NumberFamily = self.cur.fetchone()[0]

                    self.cur.execute(
                        f"""DELETE FROM anbaabram.children WHERE ChildrenNationalId = "{ChildrenNationalId}" """)
                    self.db.commit()
                    self.ResetEditChildren()
                    self.cur.execute(
                        f"""SELECT ChildrenNumber FROM anbaabram.children WHERE NumberFamily = "{NumberFamily}" """)
                    ChildrenNumber = self.cur.fetchall()
                    for row, num in enumerate(ChildrenNumber):
                        print(row, num)
                        self.cur.execute(
                            f"""UPDATE anbaabram.children SET ChildrenNumber = "{row + 1}" WHERE ChildrenNumber = "{num[0]}" """)
                        self.db.commit()
                    self.lcdNumber_5.setProperty("intValue", row + 1)
                    QMessageBox.information(self, "Info", "تم حذف الابن بنجاح")
            except:
                QMessageBox.warning(self, "Error", "الرقم القومي ليس مكون من ارقام فقط !!")
        else:
            QMessageBox.warning(self, "Error", "اكتب الرقم القومي اولا")

    # ------------------------------------------------------------- #
    #            -- Clear And Reset All Data Children --            #

    def ResetEditChildren(self):
        self.lineEdit_68.clear()
        self.lineEdit_70.clear()
        self.comboBox_10.setCurrentIndex(0)
        self.lineEdit_66.clear()
        self.lineEdit_67.clear()
        self.comboBox_16.setCurrentIndex(0)
        self.lcdNumber_5.setProperty("intValue", 0)

    # ----------------------------------------------------------------- #
    #                            -- @ Human --                          #
    # ----------------------------------------------------------------- #
    # ------------------------------------------------------------- #
    #                  -- Found All Human --                        #

    def SearchHuman(self):
        global HumanNumber
        HumanSearch = self.lineEdit_110.text()
        if HumanSearch != "":
            try:
                self.cur.execute(
                    f"""SELECT HumanName, HumanNationalId, HumanRelativeRelation, HumanSocialStatus, HumanSchoolWork, HumanMonthlyIncome, HumanFatherOfConfession FROM anbaabram.anotherhuman WHERE HumanNationalId = "{HumanSearch}" """)
                HumanData = self.cur.fetchone()
                self.lineEdit_73.setText(HumanData[0])
                self.lineEdit_72.setText(HumanData[1])
                self.lineEdit_76.setText(HumanData[2])
                self.comboBox_11.setCurrentText(HumanData[3])
                self.lineEdit_75.setText(HumanData[4])
                self.lineEdit_74.setText(HumanData[5])
                try:
                    self.cur.execute(
                        f"""INSERT INTO anbaabram.fatherofconfession (FatherOfConfession) VALUES ("{HumanData[6]}") """)
                    self.db.commit()
                    self.comboBox_23.addItem(HumanData[6])
                except:
                    pass
                self.comboBox_23.setCurrentText(HumanData[6])
                self.cur.execute(
                    f"""SELECT NumberFamily FROM anbaabram.anotherhuman WHERE HumanNationalId = "{HumanSearch}" """)
                NumFamily = self.cur.fetchone()[0]
                self.cur.execute(f"""SELECT HumanNum FROM anbaabram.anotherhuman WHERE NumberFamily = "{NumFamily}" """)
                Numbers = self.cur.fetchall()
                for H_N in Numbers:
                    HumanNumber = H_N[0]
                self.lcdNumber_4.setProperty("intValue", int(HumanNumber))
            except:
                QMessageBox.warning(self, "Error", "الرقم القومي غير صحيح !!")
        else:
            self.ResetEditHuman()

    # ------------------------------------------------------------- #
    #                 -- Save Human With New Data  --               #

    def UpdateHuman(self):
        OldId = self.lineEdit_110.text()
        HumanName = self.lineEdit_73.text()
        HumanId = self.lineEdit_72.text()
        HumanRelation = self.lineEdit_76.text()
        HumanStatusCom = self.comboBox_11.currentText()
        HumanSchool = self.lineEdit_75.text()
        HumanWork = self.lineEdit_74.text()
        HumanFatherOfConfession = self.comboBox_23.currentText()
        if OldId != "" and HumanName != "" and HumanId != "" and HumanSchool != "" and HumanWork != "":
            try:
                int(HumanId)
                try:
                    int(HumanWork)
                    if int(len(HumanId)) == 14:
                        try:
                            self.cur.execute(f"""
                                    UPDATE anbaabram.anotherhuman SET 
                                    HumanName = "{HumanName}", 
                                    HumanNationalId = "{HumanId}", 
                                    HumanRelativeRelation = "{HumanRelation}", 
                                    HumanSocialStatus = "{HumanStatusCom}", 
                                    HumanSchoolWork = "{HumanSchool}", 
                                    HumanMonthlyIncome = "{HumanWork}", 
                                    HumanFatherOfConfession = "{HumanFatherOfConfession}" 
    
                                    WHERE HumanNationalId = "{OldId}" """)
                            self.db.commit()
                            self.statusbar.showMessage("تم تحديث البيانات بنجاح بأسم -- {} --".format(HumanName))
                            self.worker.start()
                            self.worker.ClearStatusBarVariable.connect(self.StatusBar)

                            self.ResetEditHuman()
                        except:
                            QMessageBox.warning(self, "Error", "الرقم القومي الاساسي غير صحيح ...!")
                    else:
                        QMessageBox.warning(self, "Error", "يجب ان يكون الرقم القومي مكون من 14 رقم ...!")
                except:
                    QMessageBox.warning(self, "Error", "الدخل الشهري ليس مكون من ارقام فقط !!")
            except:
                QMessageBox.warning(self, "Error", "الرقم القومي ليس مكون من ارقام فقط !!")
        else:
            QMessageBox.warning(self, "Error", "يرجي اكمال البيانات لحفظها")

    # ------------------------------------------------------------- #
    #                 -- Delete All Human Data --                   #

    def DeleteHumanData(self):
        global row
        HumanNationalId = self.lineEdit_110.text()
        if HumanNationalId != "":
            try:
                int(HumanNationalId)
                self.cur.execute(
                    f"""SELECT * FROM anbaabram.anotherhuman WHERE HumanNationalId = "{HumanNationalId}" """)
                if self.cur.fetchall() == ():
                    QMessageBox.warning(self, "Error", "الرقم القومي غير مسجل بالفعل !!")
                else:
                    self.cur.execute(
                        f"""SELECT NumberFamily FROM anbaabram.anotherhuman WHERE HumanNationalId = "{HumanNationalId}" """)
                    NumberFamily = self.cur.fetchone()[0]
                    self.cur.execute(
                        f"""DELETE FROM anbaabram.anotherhuman WHERE HumanNationalId = "{HumanNationalId}" """)
                    self.db.commit()
                    self.ResetEditHuman()
                    self.cur.execute(
                        f"""SELECT HumanNum FROM anbaabram.anotherhuman WHERE NumberFamily = "{NumberFamily}" """)
                    HumanNumber = self.cur.fetchall()
                    for row, num in enumerate(HumanNumber):
                        print(row, num)
                        self.cur.execute(
                            f"""UPDATE anbaabram.anotherhuman SET HumanNum = "{row + 1}" WHERE HumanNum = "{num[0]}" """)
                        self.db.commit()
                    self.lcdNumber_4.setProperty("intValue", row + 1)
                    QMessageBox.information(self, "Info", "تم حذف الابن بنجاح")

            except:
                QMessageBox.warning(self, "Error", "الرقم القومي ليس مكون من ارقام فقط !!")
        else:
            QMessageBox.warning(self, "Error", "اكتب الرقم القومي اولا")

    # ------------------------------------------------------------- #
    #              -- Clear And Reset All Data Human --             #

    def ResetEditHuman(self):
        self.lineEdit_73.clear()
        self.lineEdit_72.clear()
        self.lineEdit_76.clear()
        self.comboBox_11.setCurrentIndex(0)
        self.lineEdit_75.clear()
        self.lineEdit_74.clear()
        self.comboBox_23.setCurrentIndex(0)
        self.lcdNumber_4.setProperty("intValue", 0)

    # ----------------------------------------------------------------- #
    #                        -- ADD HomeContent --                      #
    # ----------------------------------------------------------------- #

    def AddAddHomeContents(self):
        NumFamily = self.lineEdit_16.text()
        HomeContentsName = self.comboBox_24.currentText()
        HomeContentsNumber = self.spinBox_35.text()
        if NumFamily == "":
            QMessageBox.warning(self, "Error", "تأكد من ادخال رقم الأسرة في الصفحة الأولي ...")
        elif HomeContentsNumber == "":
            QMessageBox.warning(self, "Error", "تأكد من ادخال عدد المحتوي ...")
        elif HomeContentsName == "":
            QMessageBox.warning(self, "Error", "تأكد من تحديد المحتوي لأضافته ...")
        else:
            self.cur.execute(f"""SELECT HomeContentsName FROM anbaabram.hcontents WHERE FamilyNum = "{NumFamily}" """)
            Data = self.cur.fetchall()
            lst = []
            for data in Data:
                lst.append(data[0])
            if HomeContentsName not in lst:
                try:
                    NumFamily = int(NumFamily)
                    self.cur.execute(
                        f"""INSERT INTO anbaabram.hcontents (HomeContentsName, HomeContentsNumber, FamilyNum) 
                    VALUES ("{HomeContentsName}", "{HomeContentsNumber}", "{NumFamily}") """)
                    self.db.commit()
                    self.ShowAddHomeContents()
                except:
                    QMessageBox.warning(self, "Error", "تأكد ان رقم الاسرة مكون من ارقام فقط !")
            else:
                QMessageBox.warning(self, "Error", "المحتوي موجود بالفعل !")

    def AddRemoveHomeContents(self):
        FamilyNum = self.lineEdit_16.text()
        ComboContents = self.comboBox_24.currentText()
        if FamilyNum != "":
            self.cur.execute(f"""SELECT Contents FROM anbaabram.allhomecontents""")
            Data1 = self.cur.fetchall()
            lst1 = []
            for l in Data1:
                lst1.append(l[0])
            self.cur.execute(f"""SELECT HomeContentsName FROM anbaabram.hcontents""")
            Data2 = self.cur.fetchall()
            lst2 = []
            for l in Data2:
                lst2.append(l[0])
            if ComboContents in lst2:
                self.cur.execute(
                    f"""DELETE FROM anbaabram.hcontents WHERE HomeContentsName = "{ComboContents}" AND FamilyNum = "{FamilyNum}" """)
                self.db.commit()
                self.ShowAddHomeContents()
                self.statusbar.showMessage("تم مسح المحتوي بنجاح -- {} -- ".format(ComboContents))
                self.worker = ClearStatusBar()
                self.worker.start()
                self.worker.ClearStatusBarVariable.connect(self.StatusBar)

                self.ClearLineHomeContent()
            else:
                QMessageBox.warning(self, "Error", "المحتوي الذي ادخلته غير موجود بالفعل !")

        else:
            self.tabWidget_2.setCurrentIndex(0)
            QMessageBox.warning(self, "Error", "(رقم الأسرة) غير موجود في الصفحة الاولي  ...")

    # ----------------------------------------------------------------- #
    #                        -- Edit HomeContent --                     #
    # ----------------------------------------------------------------- #

    def EditAddHomeContents(self):
        NumFamily = self.lineEdit_57.text()
        HomeContentsName = self.comboBox_25.currentText()
        HomeContentsNumber = self.spinBox_36.text()
        if NumFamily == "":
            QMessageBox.warning(self, "Error", "تأكد من ادخال رقم الأسرة في الصفحة الأولي ...")
        elif HomeContentsNumber == "":
            QMessageBox.warning(self, "Error", "تأكد من ادخال عدد المحتوي ...")
        elif HomeContentsName == "":
            QMessageBox.warning(self, "Error", "تأكد من تحديد المحتوي لأضافته ...")
        else:
            self.cur.execute(f"""SELECT HomeContentsName FROM anbaabram.hcontents WHERE FamilyNum = "{NumFamily}" """)
            Data = self.cur.fetchall()
            lst = []
            for data in Data:
                lst.append(data[0])
            if HomeContentsName not in lst:
                try:
                    NumFamily = int(NumFamily)
                    self.cur.execute(
                        f"""INSERT INTO anbaabram.hcontents (HomeContentsName, HomeContentsNumber, FamilyNum) 
                    VALUES ("{HomeContentsName}", "{HomeContentsNumber}", "{NumFamily}") """)
                    self.db.commit()
                    self.ShowEditHomeContents()
                except:
                    QMessageBox.warning(self, "Error", "تأكد ان رقم الاسرة مكون من ارقام فقط !")
            else:
                QMessageBox.warning(self, "Error", "المحتوي موجود بالفعل !")

    def EditRemoveHomeContents(self):
        FamilyNum = self.lineEdit_57.text()
        ComboContents = self.comboBox_25.currentText()
        if FamilyNum != "":
            self.cur.execute(f"""SELECT Contents FROM anbaabram.allhomecontents""")
            Data1 = self.cur.fetchall()
            lst1 = []
            for l in Data1:
                lst1.append(l[0])
            self.cur.execute(f"""SELECT HomeContentsName FROM anbaabram.hcontents""")
            Data2 = self.cur.fetchall()
            lst2 = []
            for l in Data2:
                lst2.append(l[0])
            if ComboContents in lst2:
                self.cur.execute(
                    f"""DELETE FROM anbaabram.hcontents WHERE HomeContentsName = "{ComboContents}" AND FamilyNum = "{FamilyNum}" """)
                self.db.commit()
                self.ShowEditHomeContents()
                self.statusbar.showMessage("تم مسح المحتوي بنجاح -- {} -- ".format(ComboContents))
                self.worker = ClearStatusBar()
                self.worker.start()
                self.worker.ClearStatusBarVariable.connect(self.StatusBar)

                self.ClearLineHomeContent()
            else:
                QMessageBox.warning(self, "Error", "المحتوي الذي ادخلته غير موجود بالفعل !")

        else:
            self.tabWidget_2.setCurrentIndex(0)
            QMessageBox.warning(self, "Error", "(رقم الأسرة) غير موجود في الصفحة الاولي  ...")

    # ----------------------------------------------------------------- #

    def UpdateHomeContentComboBox(self):
        global content
        add1 = self.lineEdit_81.text()
        edit1 = self.lineEdit_82.text()
        if add1 == "" and edit1 == "":
            QMessageBox.warning(self, "Error", "برجاء ادخال اسم المحتوي لأضافته ...")
        else:
            if add1 != "":
                content = add1
            elif edit1 != "":
                content = edit1
            self.cur.execute(f"""SELECT Contents FROM anbaabram.allhomecontents""")
            Data = self.cur.fetchall()
            lst = []
            for data in Data:
                lst.append(data[0])
            if content not in lst:
                self.cur.execute(f"""INSERT INTO anbaabram.allhomecontents (Contents) VALUES ("{content}") """)
                self.db.commit()
                self.statusbar.showMessage("تم اضافة المحتوي بنجاح -- {} -- ".format(content))
                self.worker = ClearStatusBar()
                self.worker.start()
                self.worker.ClearStatusBarVariable.connect(self.StatusBar)

                self.ShowHomeContentComboBox()
                self.ClearLineHomeContent()
            else:
                QMessageBox.warning(self, "Error", "المحتوي الذي ادخلته موجود بالفعل !")

    def DeleteHomeContent(self):
        global content
        add1 = self.lineEdit_81.text()
        edit1 = self.lineEdit_82.text()
        if add1 == "" and edit1 == "":
            QMessageBox.warning(self, "Error", "برجاء ادخال اسم المحتوي لحذفه ...")
        else:
            if add1 != "":
                content = add1
            elif edit1 != "":
                content = edit1
            self.cur.execute(f"""SELECT Contents FROM anbaabram.allhomecontents""")
            Data = self.cur.fetchall()
            lst = []
            for data in Data:
                lst.append(data[0])
            if content in lst:
                self.cur.execute(f"""DELETE FROM anbaabram.allhomecontents WHERE Contents = "{content}" """)
                self.db.commit()
                lst.remove(content)
                self.statusbar.showMessage("تم مسح -- {} -- بنجاح".format(content))
                self.worker = ClearStatusBar()
                self.worker.start()
                self.worker.ClearStatusBarVariable.connect(self.StatusBar)

                self.ShowHomeContentComboBox()
                self.ClearLineHomeContent()
            else:
                QMessageBox.warning(self, "Error", "المحتوي الذي ادخلته موجود بالفعل !")

    def ClearLineHomeContent(self):
        self.lineEdit_81.setText("")
        self.lineEdit_82.setText("")

    # ------------------------------------------------------------------------------------------
    def ShowAddHomeContents(self):
        NumFamily = self.lineEdit_16.text()
        self.tableWidget.setRowCount(0)
        self.tableWidget.insertRow(0)
        self.cur.execute(
            f"""SELECT HomeContentsName, HomeContentsNumber FROM anbaabram.hcontents WHERE FamilyNum = "{NumFamily}" """)
        Data = self.cur.fetchall()
        for row, form in enumerate(Data):
            for col, item in enumerate(form):
                self.tableWidget.setItem(row, col, QTableWidgetItem(str(item)))
                col += 1
            row_position = self.tableWidget.rowCount()
            self.tableWidget.insertRow(row_position)

    def ShowEditHomeContents(self):
        NumFamily = self.lineEdit_57.text()
        self.tableWidget_2.setRowCount(0)
        self.tableWidget_2.insertRow(0)
        self.cur.execute(
            f"""SELECT HomeContentsName, HomeContentsNumber FROM anbaabram.hcontents WHERE FamilyNum = "{NumFamily}" """)
        Data = self.cur.fetchall()
        for row, form in enumerate(Data):
            for col, item in enumerate(form):
                self.tableWidget_2.setItem(row, col, QTableWidgetItem(str(item)))
                col += 1
            row_position = self.tableWidget_2.rowCount()
            self.tableWidget_2.insertRow(row_position)

    # ------------------------------------------------------------------------------------------
    def ShowHomeContentComboBox(self):
        self.comboBox_24.clear()
        self.comboBox_25.clear()
        self.cur.execute("""SELECT * FROM anbaabram.allhomecontents""")
        names = self.cur.fetchall()
        for name in names:
            self.comboBox_24.addItem(name[1])
            self.comboBox_25.addItem(name[1])

    # ----------------------------------------------------------------- #
    #                       -- FatherOfConfession --                    #
    # ----------------------------------------------------------------- #

    def AddFatherOfConfession(self):
        global name
        add1 = self.lineEdit_80.text()
        add2 = self.lineEdit_27.text()
        add3 = self.lineEdit_33.text()
        edit1 = self.lineEdit_79.text()
        edit2 = self.lineEdit_69.text()
        edit3 = self.lineEdit_71.text()
        if add1 == "" and add2 == "" and add3 == "" and edit1 == "" and edit2 == "" and edit3 == "":
            QMessageBox.warning(self, "Error", "برجاء ادخال اسم لأضافته ...")
        else:
            if add1 != "":
                name = add1
            elif add2 != "":
                name = add2
            elif add3 != "":
                name = add3
            elif edit1 != "":
                name = edit1
            elif edit2 != "":
                name = edit2
            elif edit3 != "":
                name = edit3
            self.cur.execute("SELECT FatherOfConfession FROM anbaabram.fatherofconfession")
            names = self.cur.fetchall()
            n = []
            for Name in names:
                n.append(Name[0])
            if name not in n:
                self.cur.execute(
                    f"""INSERT INTO anbaabram.fatherofconfession (FatherOfConfession) VALUES ("{name}") """)
                self.db.commit()
                self.ShowFatherOfConfession()
                self.statusbar.showMessage("تم اضافة اسم اب الاعتراف بنجاح -- {} --".format(name))

                self.worker = ClearStatusBar()
                self.worker = ClearStatusBar()
                self.worker.start()
                self.worker.ClearStatusBarVariable.connect(self.StatusBar)
                self.ClearLineFatherOfConfession()
            else:
                QMessageBox.warning(self, "Error", "الأسم موجود بالفعل -- {} --".format(name))

    def DeleteFatherOfConfession(self):
        global name
        add1 = self.lineEdit_80.text()
        add2 = self.lineEdit_27.text()
        add3 = self.lineEdit_33.text()
        edit1 = self.lineEdit_79.text()
        edit2 = self.lineEdit_69.text()
        edit3 = self.lineEdit_71.text()
        if add1 == "" and add2 == "" and add3 == "" and edit1 == "" and edit2 == "" and edit3 == "":
            QMessageBox.warning(self, "Error", "برجاء ادخال اسم لحذفه ...")
        else:
            if add1 != "":
                name = add1
            elif add2 != "":
                name = add2
            elif add3 != "":
                name = add3
            elif edit1 != "":
                name = edit1
            elif edit2 != "":
                name = edit2
            elif edit3 != "":
                name = edit3
            self.cur.execute("SELECT FatherOfConfession FROM anbaabram.fatherofconfession")
            names = self.cur.fetchall()
            n = []
            for Name in names:
                n.append(Name[0])
            if name in n:
                self.cur.execute(f"""DELETE FROM anbaabram.fatherofconfession WHERE FatherOfConfession = "{name}" """)
                self.db.commit()
                n.remove(name)
                self.statusbar.showMessage("بنجاح -- {} -- تم مسح".format(name))
                self.ShowFatherOfConfession()
                self.worker = ClearStatusBar()
                self.worker.start()
                self.worker.ClearStatusBarVariable.connect(self.StatusBar)

                self.ClearLineFatherOfConfession()
            else:
                QMessageBox.warning(self, "Error", "الأسم غير موجود بالفعل -- {} --".format(name))

    def ClearLineFatherOfConfession(self):
        self.lineEdit_80.clear()
        self.lineEdit_27.clear()
        self.lineEdit_33.clear()
        self.lineEdit_79.clear()
        self.lineEdit_69.clear()
        self.lineEdit_71.clear()

    # ----------------------------------------------------------------- #

    def ShowFatherOfConfession(self):
        self.comboBox_20.clear()
        self.comboBox_19.clear()
        self.comboBox_15.clear()
        self.comboBox_22.clear()
        self.comboBox_16.clear()
        self.comboBox_17.clear()
        self.comboBox_18.clear()
        self.comboBox_23.clear()
        self.cur.execute("""SELECT FatherOfConfession FROM fatherofconfession""")
        names = self.cur.fetchall()
        for name in names:
            self.comboBox_20.addItem(name[0])
            self.comboBox_19.addItem(name[0])
            self.comboBox_15.addItem(name[0])
            self.comboBox_22.addItem(name[0])
            self.comboBox_16.addItem(name[0])
            self.comboBox_17.addItem(name[0])
            self.comboBox_18.addItem(name[0])
            self.comboBox_23.addItem(name[0])

    # ----------------------------------------------------------------- #

    # ----------------------------------------------------------------- #
    #                          -- Governorate --                        #
    # ----------------------------------------------------------------- #

    def AddGovernorate(self):
        global name
        add1 = self.lineEdit_17.text()
        edit1 = self.lineEdit_104.text()
        if add1 == "" and edit1 == "":
            QMessageBox.warning(self, "Error", "برجاء ادخال اسم لأضافته ...")
        else:
            if add1 != "":
                name = add1
            elif edit1 != "":
                name = edit1
            self.cur.execute("SELECT GovernorateName FROM anbaabram.governorate")
            names = self.cur.fetchall()
            n = []
            for Name in names:
                n.append(Name[0])
            if name not in n:
                self.cur.execute(f"""INSERT INTO anbaabram.governorate (GovernorateName) VALUES ("{name}") """)
                self.db.commit()
                self.ShowGovernorate()
                self.statusbar.showMessage("تم اضافة اسم المحافظة بنجاح -- {} --".format(name))
                self.worker = ClearStatusBar()
                self.worker.start()
                self.worker.ClearStatusBarVariable.connect(self.StatusBar)

                self.ClearLineGovernorate()
            else:
                QMessageBox.warning(self, "Error", "الأسم موجود بالفعل -- {} --".format(name))

    def DeleteGovernorate(self):
        global name
        add1 = self.lineEdit_17.text()
        edit1 = self.lineEdit_104.text()
        if add1 == "" and edit1 == "":
            QMessageBox.warning(self, "Error", "برجاء ادخال اسم لحذفه ...")
        else:
            if add1 != "":
                name = add1
            elif edit1 != "":
                name = edit1
            self.cur.execute("SELECT GovernorateName FROM anbaabram.governorate")
            names = self.cur.fetchall()
            n = []
            for Name in names:
                n.append(Name[0])
            if name in n:
                self.cur.execute(f"""DELETE FROM anbaabram.governorate WHERE GovernorateName = "{name}" """)
                self.db.commit()
                n.remove(name)
                self.statusbar.showMessage("بنجاح -- {} -- تم مسح".format(name))
                self.worker = ClearStatusBar()
                self.worker.start()
                self.worker.ClearStatusBarVariable.connect(self.StatusBar)

                self.ShowGovernorate()
                self.ClearLineGovernorate()
            else:
                QMessageBox.warning(self, "Error", "الأسم غير موجود بالفعل -- {} --".format(name))

    def ClearLineGovernorate(self):
        self.lineEdit_17.clear()
        self.lineEdit_104.clear()

    # ----------------------------------------------------------------- #

    def ShowGovernorate(self):
        self.comboBox.clear()
        self.comboBox_9.clear()
        self.cur.execute(f"""SELECT GovernorateName FROM anbaabram.governorate""")
        Data = self.cur.fetchall()
        for data in Data:
            self.comboBox.addItem(data[0])
            self.comboBox_9.addItem(data[0])

    # ----------------------------------------------------------------- #

    # ----------------------------------------------------------------- #
    #                            -- District --                         #
    # ----------------------------------------------------------------- #

    def AddDistrict(self):
        global name
        add1 = self.lineEdit_18.text()
        edit1 = self.lineEdit_107.text()

        if add1 == "" and edit1 == "":
            QMessageBox.warning(self, "Error", "برجاء ادخال اسم لأضافته ...")
        else:
            if add1 != "":
                name = add1
            elif edit1 != "":
                name = edit1
            self.cur.execute("SELECT DistrictName FROM anbaabram.district")
            names = self.cur.fetchall()
            n = []
            for Name in names:
                n.append(Name[0])
            if name not in n:
                self.cur.execute(f"""INSERT INTO anbaabram.district (DistrictName) VALUES ("{name}") """)
                self.db.commit()
                self.ShowDistrict()
                self.statusbar.showMessage("تم اضافة اسم الحي بنجاح -- {} --".format(name))
                self.worker = ClearStatusBar()
                self.worker.start()
                self.worker.ClearStatusBarVariable.connect(self.StatusBar)

                self.ClearLineDistrict()
            else:
                QMessageBox.warning(self, "Error", "الأسم موجود بالفعل -- {} --".format(name))

    def DeleteDistrict(self):
        global name
        add1 = self.lineEdit_18.text()
        edit1 = self.lineEdit_107.text()

        if add1 == "" and edit1 == "":
            QMessageBox.warning(self, "Error", "برجاء ادخال اسم لأضافته ...")
        else:
            if add1 != "":
                name = add1
            elif edit1 != "":
                name = edit1
            self.cur.execute("SELECT DistrictName FROM anbaabram.district")
            names = self.cur.fetchall()
            n = []
            for Name in names:
                n.append(Name[0])
            if name in n:
                self.cur.execute(f"""DELETE FROM anbaabram.district WHERE DistrictName = "{name}" """)
                self.db.commit()
                n.remove(name)
                self.statusbar.showMessage("بنجاح -- {} -- تم مسح".format(name))
                self.worker = ClearStatusBar()
                self.worker.start()
                self.worker.ClearStatusBarVariable.connect(self.StatusBar)

                self.ShowDistrict()
                self.ClearLineDistrict()
            else:
                QMessageBox.warning(self, "Error", "الأسم غير موجود بالفعل -- {} --".format(name))

    def ClearLineDistrict(self):
        self.lineEdit_18.clear()
        self.lineEdit_107.clear()

    # ----------------------------------------------------------------- #

    def ShowDistrict(self):
        self.comboBox_2.clear()
        self.comboBox_8.clear()
        self.cur.execute(f"""SELECT DistrictName FROM anbaabram.district""")
        Data = self.cur.fetchall()
        for data in Data:
            self.comboBox_2.addItem(data[0])
            self.comboBox_8.addItem(data[0])

    # ----------------------------------------------------------------- #

    # ----------------------------------------------------------------- #
    #                          -- Address --                            #
    # ----------------------------------------------------------------- #

    def AddAddress(self):
        global name
        add1 = self.lineEdit_22.text()
        edit1 = self.lineEdit_108.text()
        if add1 == "" and edit1 == "":
            QMessageBox.warning(self, "Error", "برجاء ادخال اسم لأضافته ...")
        else:
            if add1 != "":
                name = add1
            elif edit1 != "":
                name = edit1
            self.cur.execute("SELECT addressname FROM anbaabram.address")
            names = self.cur.fetchall()
            n = []
            for Name in names:
                n.append(Name[0])
            if name not in n:
                self.cur.execute(f"""INSERT INTO anbaabram.address (addressname) VALUES ("{name}") """)
                self.db.commit()
                self.ShowAddress()
                self.statusbar.showMessage("تم اضافة العنوان بنجاح -- {} --".format(name))
                self.worker = ClearStatusBar()
                self.worker.start()
                self.worker.ClearStatusBarVariable.connect(self.StatusBar)

                self.ClearLineAddress()
            else:
                QMessageBox.warning(self, "Error", "الأسم موجود بالفعل -- {} --".format(name))

    def DeleteAddress(self):
        global name
        add1 = self.lineEdit_22.text()
        edit1 = self.lineEdit_108.text()
        if add1 == "" and edit1 == "":
            QMessageBox.warning(self, "Error", "برجاء ادخال اسم لحذفه ...")
        else:
            if add1 != "":
                name = add1
            elif edit1 != "":
                name = edit1
            self.cur.execute("SELECT addressname FROM anbaabram.address")
            names = self.cur.fetchall()
            n = []
            for Name in names:
                n.append(Name[0])
            if name in n:
                self.cur.execute(f"""DELETE FROM anbaabram.address WHERE addressname = "{name}" """)
                self.db.commit()
                n.remove(name)
                self.statusbar.showMessage("بنجاح -- {} -- تم مسح".format(name))
                self.worker = ClearStatusBar()
                self.worker.start()
                self.worker.ClearStatusBarVariable.connect(self.StatusBar)

                self.ShowAddress()
                self.ClearLineAddress()
            else:
                QMessageBox.warning(self, "Error", "الأسم غير موجود بالفعل -- {} --".format(name))

    # ----------------------------------------------------------------- #

    def ClearLineAddress(self):
        self.lineEdit_22.clear()
        self.lineEdit_108.clear()

    # ----------------------------------------------------------------- #

    def ShowAddress(self):
        self.comboBox_3.clear()
        self.comboBox_7.clear()
        self.cur.execute(f"""SELECT addressname FROM anbaabram.address""")
        Data = self.cur.fetchall()
        for data in Data:
            self.comboBox_3.addItem(data[0])
            self.comboBox_7.addItem(data[0])

    # ----------------------------------------------------------------- #

    # ----------------------------------------------------------------- #
    #                          -- Source --                             #
    # ----------------------------------------------------------------- #
    def AddSource(self):
        global name
        add1 = self.lineEdit_36.text()
        edit1 = self.lineEdit_77.text()
        if add1 == "" and edit1 == "":
            QMessageBox.warning(self, "Error", "برجاء ادخال اسم لأضافته ...")
        else:
            if add1 != "":
                name = add1
            elif edit1 != "":
                name = edit1
            self.cur.execute("SELECT SourceName FROM anbaabram.source")
            names = self.cur.fetchall()
            n = []
            for Name in names:
                n.append(Name[0])
            if name not in n:
                self.cur.execute(f"""INSERT INTO anbaabram.source (SourceName) VALUES ("{name}") """)
                self.db.commit()
                self.ShowSource()
                self.statusbar.showMessage("تم اضافة اسم الحي بنجاح -- {} --".format(name))
                self.worker = ClearStatusBar()
                self.worker.start()
                self.worker.ClearStatusBarVariable.connect(self.StatusBar)

                self.ClearLineSource()
            else:
                QMessageBox.warning(self, "Error", "الأسم موجود بالفعل -- {} --".format(name))

    def DeleteSource(self):
        global name
        add1 = self.lineEdit_36.text()
        edit1 = self.lineEdit_77.text()
        if add1 == "" and edit1 == "":
            QMessageBox.warning(self, "Error", "برجاء ادخال اسم لحذفه ...")
        else:
            if add1 != "":
                name = add1
            elif edit1 != "":
                name = edit1
            self.cur.execute("SELECT SourceName FROM anbaabram.source")
            names = self.cur.fetchall()
            n = []
            for Name in names:
                n.append(Name[0])
            if name in n:
                self.cur.execute(f"""DELETE FROM anbaabram.source WHERE SourceName = "{name}" """)
                self.db.commit()
                n.remove(name)
                self.statusbar.showMessage("بنجاح -- {} -- تم مسح".format(name))
                self.worker = ClearStatusBar()
                self.worker.start()
                self.worker.ClearStatusBarVariable.connect(self.StatusBar)

                self.ShowSource()
                self.ClearLineSource()
            else:
                QMessageBox.warning(self, "Error", "الأسم غير موجود بالفعل -- {} --".format(name))

    def ClearLineSource(self):
        self.lineEdit_36.clear()
        self.lineEdit_77.clear()

    def ShowSource(self):
        self.comboBox_6.clear()
        self.comboBox_12.clear()
        self.cur.execute(f"""SELECT SourceName FROM anbaabram.source""")
        Data = self.cur.fetchall()
        for data in Data:
            self.comboBox_6.addItem(data[0])
            self.comboBox_12.addItem(data[0])

    # ----------------------------------------------------------------- #
    #                        -- Open Tabs --                            #
    # ----------------------------------------------------------------- #

    def MoveLogo(self):
        l = QPropertyAnimation(self.label_30, b"geometry")
        l.setDuration(500)
        l.setStartValue(QRect(330, 100, 540, 150))
        l.setEndValue(QRect(330, 80, 540, 150))
        l.start()
        self.l = l

    def BG(self):
        bg = QPropertyAnimation(self.label_6, b"geometry")
        bg.setDuration(8500)
        bg.setStartValue(QRect(-1450, -600, 2800, 1460))
        bg.setEndValue(QRect(-200, -30, 2800, 1460))
        bg.start()
        self.bg = bg

    def BG1(self):
        bg1 = QPropertyAnimation(self.label_7, b"geometry")
        bg1.setDuration(10000)
        bg1.setStartValue(QRect(-650, -80, 1690, 970))
        bg1.setEndValue(QRect(-10, -80, 1690, 970))
        bg1.start()
        self.bg1 = bg1

    def OpenShowTab(self):
        self.tabWidget.setCurrentIndex(3)

    def OpenAdds(self):
        self.tabWidget.setCurrentIndex(4)
        self.tabWidget_2.setCurrentIndex(0)
        self.tabWidget_3.setCurrentIndex(0)
        self.tabWidget_4.setCurrentIndex(0)

        self.ClearLineHomeContent()

    def OpenEditOrDeleteTab(self):
        self.tabWidget.setCurrentIndex(5)
        self.tabWidget_6.setCurrentIndex(0)
        self.tabWidget_5.setCurrentIndex(0)
        self.tabWidget_7.setCurrentIndex(0)

        self.ClearLineHomeContent()

    # ---------------------------------------------------------------------------------------------------------------------------------- #

    def ClearLinesFunction(self):
        self.ClearLineFatherOfConfession()
        self.ClearLineGovernorate()
        self.ClearLineDistrict()
        self.ClearLineAddress()
        self.ClearLineSource()

    # ----------------------------------------------------------------- #
    #                        -- Add Next --                             #
    # ----------------------------------------------------------------- #

    def AddNext0(self):
        NumFamily = self.lineEdit_16.text()
        if NumFamily != "":
            try:
                int(NumFamily)
                self.Page1HusbandAndWife()
                try:
                    self.UpdateNumberChildren()
                    self.UpdateNumberOtherHuman()
                except:
                    pass
            except:
                QMessageBox.warning(self, "Error", "تأكد من رقم الاسرة اولا ...")
        else:
            QMessageBox.warning(self, "Error", "يرجي التأكد من رقم الاسرة ...")

    def AddNext1(self):
        self.tabWidget_2.setCurrentIndex(2)
        self.tabWidget_3.setCurrentIndex(0)
        self.ClearLinesFunction()

    def AddNext2_1(self):
        check = self.checkBox.checkState().__str__()
        if check == '2':
            self.tabWidget_3.setCurrentIndex(1)
        else:
            self.tabWidget_2.setCurrentIndex(3)
            self.tabWidget_4.setCurrentIndex(0)
        self.ClearLinesFunction()
        self.ShowAddHomeContents()

    def AddNext2_2(self):
        self.tabWidget_2.setCurrentIndex(3)
        self.tabWidget_4.setCurrentIndex(0)
        self.ClearLinesFunction()

    def AddNext3_1(self):
        check = self.checkBox_2.checkState().__str__()
        if check == '2':
            self.tabWidget_4.setCurrentIndex(1)
        else:
            self.tabWidget_2.setCurrentIndex(4)
        self.ClearLinesFunction()

    def AddNext3_2(self):
        self.tabWidget_2.setCurrentIndex(4)
        self.ClearLinesFunction()

    def AddNext4(self):
        self.tabWidget_2.setCurrentIndex(5)
        self.ClearLinesFunction()

    # ---------------------------------------------------------------------------------------------------------------------------------- #

    # ----------------------------------------------------------------- #
    #                        -- Add Back --                             #
    # ----------------------------------------------------------------- #

    def AddBack1(self):
        self.tabWidget_2.setCurrentIndex(0)

    def AddBack2_1(self):
        self.tabWidget_2.setCurrentIndex(1)

    def AddBack2_2(self):
        self.tabWidget_3.setCurrentIndex(0)

    def AddBack3_1(self):
        check = self.checkBox.checkState().__str__()
        if check == '2':
            self.tabWidget_2.setCurrentIndex(2)
            self.tabWidget_3.setCurrentIndex(1)
        else:
            self.tabWidget_2.setCurrentIndex(2)
            self.tabWidget_3.setCurrentIndex(0)

    def AddBack3_2(self):
        self.tabWidget_4.setCurrentIndex(0)

    def AddBack4(self):
        check = self.checkBox_2.checkState().__str__()
        if check == '2':
            self.tabWidget_2.setCurrentIndex(3)
            self.tabWidget_4.setCurrentIndex(1)
        else:
            self.tabWidget_2.setCurrentIndex(3)
            self.tabWidget_4.setCurrentIndex(0)

    def AddBack5(self):
        self.tabWidget_2.setCurrentIndex(4)

    # ---------------------------------------------------------------------------------------------------------------------------------- #

    # ----------------------------------------------------------------- #
    #                        -- Edit Next --                            #
    # ----------------------------------------------------------------- #

    def EditNext0(self):
        self.tabWidget_5.setCurrentIndex(1)
        self.ShowEditHomeContents()

    def EditNext1(self):
        self.tabWidget_5.setCurrentIndex(2)
        self.tabWidget_6.setCurrentIndex(0)
        self.ClearLinesFunction()

    def EditNext2_1(self):
        check = self.checkBox_7.checkState().__str__()
        if check == '2':
            self.tabWidget_6.setCurrentIndex(1)
        else:
            self.tabWidget_5.setCurrentIndex(3)
            self.tabWidget_7.setCurrentIndex(0)
        self.ClearLinesFunction()

    def EditNext2_2(self):
        self.tabWidget_5.setCurrentIndex(3)
        self.tabWidget_7.setCurrentIndex(0)
        self.ClearLinesFunction()

    def EditNext3_1(self):
        check = self.checkBox_8.checkState().__str__()
        if check == '2':
            self.tabWidget_7.setCurrentIndex(1)
        else:
            self.tabWidget_5.setCurrentIndex(4)
        self.ClearLinesFunction()

    def EditNext3_2(self):
        self.tabWidget_5.setCurrentIndex(4)
        self.ClearLinesFunction()

    def EditNext4(self):
        self.tabWidget_5.setCurrentIndex(5)
        self.ClearLinesFunction()

    # ---------------------------------------------------------------------------------------------------------------------------------- #

    # ----------------------------------------------------------------- #
    #                        -- Edit Back --                            #
    # ----------------------------------------------------------------- #

    def EditBack1(self):
        self.tabWidget_5.setCurrentIndex(0)

    def EditBack2_1(self):
        self.tabWidget_5.setCurrentIndex(1)

    def EditBack2_2(self):
        self.tabWidget_6.setCurrentIndex(0)

    def EditBack3_1(self):
        check = self.checkBox_7.checkState().__str__()
        if check == '2':
            self.tabWidget_5.setCurrentIndex(2)
            self.tabWidget_6.setCurrentIndex(1)
        else:
            self.tabWidget_5.setCurrentIndex(2)
            self.tabWidget_6.setCurrentIndex(0)

    def EditBack3_2(self):
        self.tabWidget_7.setCurrentIndex(0)

    def EditBack4(self):
        check = self.checkBox_8.checkState().__str__()
        if check == '2':
            self.tabWidget_5.setCurrentIndex(3)
            self.tabWidget_7.setCurrentIndex(1)
        else:
            self.tabWidget_5.setCurrentIndex(3)
            self.tabWidget_7.setCurrentIndex(0)

    def EditBack5(self):
        self.tabWidget_5.setCurrentIndex(4)

    # ---------------------------------------------------------------- #
    #                                                                  #
    #                         ----- End ----                           #
    #                                                                  #
    # ---------------------------------------------------------------- #

    # ---------------------------------------------------------------- #
    #                                                                  #
    #                          -- Excel --                             #
    #                                                                  #
    # ---------------------------------------------------------------- #


class Process(QThread):
    T = pyqtSignal(list)

    def run(self):
        text = "        قاعدة بيانات خدمة الأنبا ابرام"
        litter = ""
        lst = []
        for t in text:
            sleep(.07)
            litter = litter + t
            lst.append(litter)
            self.T.emit(lst)


class ProgressBar(QThread):
    prog = pyqtSignal(int)

    def run(self):
        for value in range(140):
            self.prog.emit(value - 40)
            sleep(.007)
        sleep(1)


class ClearStatusBar(QThread):
    ClearStatusBarVariable = pyqtSignal(str)

    def run(self):
        sleep(5)
        self.ClearStatusBarVariable.emit('')


def main():
    app = QApplication([])
    win = Main()
    win.show()
    app.exec_()


if __name__ == '__main__':
    main()
